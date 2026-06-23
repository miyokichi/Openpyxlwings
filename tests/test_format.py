from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import Border, Side

from openpyxlwings import ExcelFormat, ExcelWorkbook
from openpyxlwings.exceptions import FormatDefinitionError, FormatMatchError


def apply_grid(sheet, start_row: int, start_column: int, end_row: int, end_column: int) -> None:
    thin = Side(style="thin")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for row in range(start_row, end_row + 1):
        for column in range(start_column, end_column + 1):
            sheet.cell(row=row, column=column).border = border


def make_format_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "amount_table"
    sheet["B2"] = "header1"
    sheet["C2"] = "header2"
    sheet["D2"] = '{{columns[].header | contains("amount")}}'
    sheet["B3"] = "{{rows[].header1}}"
    sheet["C3"] = "{{rows[].header2}}"
    sheet["D3"] = "{{rows[].amounts[]:float}}"
    apply_grid(sheet, 2, 2, 3, 4)
    workbook.save(path)


def write_amount_table(sheet, start_row: int, start_column: int, *, formula: bool = False) -> None:
    values = [
        ["header1", "header2", "amount", "amount forecast", "amount final"],
        ["row1", "sub1", 100, 120, 140],
        ["row2", "sub2", 200, 220, 240],
        ["row3", "sub3", 300, 320, 340],
    ]
    for row_offset, row_values in enumerate(values):
        for column_offset, value in enumerate(row_values):
            sheet.cell(start_row + row_offset, start_column + column_offset, value=value)
    if formula:
        sheet.cell(start_row + 1, start_column + 2, value="=50+50")
    apply_grid(
        sheet,
        start_row,
        start_column,
        start_row + len(values) - 1,
        start_column + len(values[0]) - 1,
    )


def make_target_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    write_amount_table(sheet, 2, 2, formula=True)
    write_amount_table(sheet, 9, 3)
    other = workbook.create_sheet("Other")
    write_amount_table(other, 4, 1)
    workbook.save(path)


def test_excel_format_loads_one_pattern_per_sheet(tmp_path: Path) -> None:
    path = tmp_path / "formats.xlsx"
    make_format_workbook(path)

    formats = ExcelFormat.load(path)
    pattern = formats["amount_table"]

    assert list(formats) == ["amount_table"]
    assert pattern.source_range == "B2:D3"
    assert pattern.row_repeat_index == 1
    assert pattern.column_repeat_index == 2


def test_placeholder_equals_constraint_and_types() -> None:
    from openpyxlwings import Placeholder

    placeholder = Placeholder.parse('{{value:int | equals("12")}}')

    assert placeholder is not None
    assert placeholder.convert("12", source="Sheet1!A1") == 12


def test_extract_scalar_placeholders(tmp_path: Path) -> None:
    format_path = tmp_path / "formats.xlsx"
    format_book = Workbook()
    pattern_sheet = format_book.active
    pattern_sheet.title = "report_info"
    pattern_sheet["A1"] = "report_title"
    pattern_sheet["B1"] = "{{title}}"
    pattern_sheet["A2"] = "report_date"
    pattern_sheet["B2"] = "{{report_date:date}}"
    apply_grid(pattern_sheet, 1, 1, 2, 2)
    format_book.save(format_path)

    target_path = tmp_path / "input.xlsx"
    target_book = Workbook()
    target_sheet = target_book.active
    target_sheet.title = "Info"
    target_sheet["C3"] = "report_title"
    target_sheet["D3"] = "Monthly report"
    target_sheet["C4"] = "report_date"
    target_sheet["D4"] = date(2026, 6, 24)
    apply_grid(target_sheet, 3, 3, 4, 4)
    target_book.save(target_path)

    pattern = ExcelFormat.load(format_path)["report_info"]
    with ExcelWorkbook(target_path) as reader:
        matches = reader.extract(pattern)

    assert matches[0].data == {
        "title": "Monthly report",
        "report_date": date(2026, 6, 24),
    }
    assert matches[0].formulas == {"title": None, "report_date": None}
    assert matches[0].source_cells == {
        "title": "Info!D3",
        "report_date": "Info!D4",
    }


def test_extract_converts_variable_rows_and_columns(tmp_path: Path) -> None:
    format_path = tmp_path / "formats.xlsx"
    target_path = tmp_path / "input.xlsx"
    make_format_workbook(format_path)
    make_target_workbook(target_path)
    pattern = ExcelFormat.load(format_path)["amount_table"]

    with ExcelWorkbook(target_path) as workbook:
        matches = workbook.extract(pattern, sheets=["Data"])

    assert [match.range for match in matches] == ["B2:F5", "C9:G12"]
    first = matches[0]
    assert first.data["columns"] == [
        {"header": "amount"},
        {"header": "amount forecast"},
        {"header": "amount final"},
    ]
    assert first.data["rows"] == [
        {"header1": "row1", "header2": "sub1", "amounts": [None, 120.0, 140.0]},
        {"header1": "row2", "header2": "sub2", "amounts": [200.0, 220.0, 240.0]},
        {"header1": "row3", "header2": "sub3", "amounts": [300.0, 320.0, 340.0]},
    ]
    assert first.formulas["rows"][0]["amounts"] == ["=50+50", None, None]
    assert first.source_cells["rows"][0]["amounts"] == ["Data!D3", "Data!E3", "Data!F3"]


def test_extract_can_limit_sheets_and_ranges(tmp_path: Path) -> None:
    format_path = tmp_path / "formats.xlsx"
    target_path = tmp_path / "input.xlsx"
    make_format_workbook(format_path)
    make_target_workbook(target_path)
    pattern = ExcelFormat.load(format_path)["amount_table"]

    with ExcelWorkbook(target_path) as workbook:
        matches = workbook.extract(pattern, ranges={"Data": "A1:H6"})

    assert [(match.sheet, match.range) for match in matches] == [("Data", "B2:F5")]


def test_extract_requires_exact_literals_and_constraints(tmp_path: Path) -> None:
    format_path = tmp_path / "formats.xlsx"
    target_path = tmp_path / "input.xlsx"
    make_format_workbook(format_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    write_amount_table(sheet, 2, 2)
    sheet["B2"] = "Header1"
    workbook.save(target_path)
    pattern = ExcelFormat.load(format_path)["amount_table"]

    with ExcelWorkbook(target_path) as reader:
        assert reader.extract(pattern) == []


def test_pattern_without_literal_requires_range(tmp_path: Path) -> None:
    format_path = tmp_path / "formats.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "anonymous"
    sheet["A1"] = "{{columns[].header}}"
    sheet["A2"] = "{{rows[].values[]}}"
    apply_grid(sheet, 1, 1, 2, 1)
    workbook.save(format_path)
    target_path = tmp_path / "input.xlsx"
    make_target_workbook(target_path)
    pattern = ExcelFormat.load(format_path)["anonymous"]

    with ExcelWorkbook(target_path) as reader:
        with pytest.raises(FormatMatchError):
            reader.extract(pattern)


def test_invalid_repeat_layout_is_rejected(tmp_path: Path) -> None:
    format_path = tmp_path / "formats.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "invalid"
    sheet["A1"] = "{{rows[].name}}"
    sheet["A2"] = "footer"
    apply_grid(sheet, 1, 1, 2, 1)
    workbook.save(format_path)

    with pytest.raises(FormatDefinitionError):
        ExcelFormat.load(format_path)


def test_format_pattern_requires_complete_borders(tmp_path: Path) -> None:
    format_path = tmp_path / "formats.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "invalid_border"
    sheet["A1"] = "header1"
    sheet["B1"] = "header2"
    workbook.save(format_path)

    with pytest.raises(FormatDefinitionError):
        ExcelFormat.load(format_path)


def test_placeholder_must_match_its_repeat_axes(tmp_path: Path) -> None:
    format_path = tmp_path / "formats.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "invalid_axis"
    sheet["A1"] = "header"
    sheet["B1"] = "{{columns[].header}}"
    sheet["A2"] = "{{rows[].name}}"
    sheet["B2"] = "{{rows[].wrong}}"
    apply_grid(sheet, 1, 1, 2, 2)
    workbook.save(format_path)

    with pytest.raises(FormatDefinitionError):
        ExcelFormat.load(format_path)
