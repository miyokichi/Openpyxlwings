"""Tests for the WritePlan / ExcelWorkbook.apply sample code.

These exercise the README examples without a real Excel install by swapping
the xlwings write session for a recorder. The plan building, validation, and
dispatch happen in pure Python, so they run anywhere.
"""

from pathlib import Path

import pytest
from openpyxl import Workbook

from openpyxlwings import BorderTable, ExcelWorkbook, WritePlan


def make_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Summary"
    sheet["A1"] = "title"
    workbook.create_sheet("Data")
    workbook.save(path)


class RecordingWriter:
    """Stand-in for ``_XlwingsWriteSession`` that records calls instead of
    driving Excel."""

    def __init__(self) -> None:
        self.calls: list[tuple] = []
        self.saved = 0

    def write_values(self, sheet, cell, values, *, expand=False) -> None:
        self.calls.append(("write_values", sheet, cell, values, expand))

    def clear_contents(self, sheet, address) -> None:
        self.calls.append(("clear_contents", sheet, address))

    def apply_bordered_table(
        self,
        sheet,
        *,
        partial,
        start_row,
        start_column,
        header_rows,
        columns,
        end_row,
        end_column,
        insertions,
        partial_axis=None,
        source_rows=None,
    ) -> None:
        self.calls.append(
            (
                "apply_bordered_table",
                sheet,
                partial,
                start_row,
                start_column,
                header_rows,
                [(source, list(values)) for source, values in columns],
                end_row,
                end_column,
                tuple(insertions),
            )
        )

    def save(self, path=None) -> None:
        self.saved += 1

    def close(self, *, save=True) -> None:
        self.closed_with_save = save


def test_plan_records_operations_in_order() -> None:
    plan = WritePlan()
    plan.write_values("Summary", "B2", "更新済み")
    plan.clear_contents("Data", "A2:F1000")

    assert len(plan) == 2
    ops = list(plan)
    assert ops[0].cell == "B2"
    assert ops[1].address == "A2:F1000"


def test_at_methods_normalize_to_addresses() -> None:
    plan = WritePlan()
    plan.write_values_at("Data", row=2, column=1, values=[[1, 2], [3, 4]])
    plan.clear_contents_at("Data", start_row=10, start_column=1, end_row=100, end_column=6)

    ops = list(plan)
    assert ops[0].cell == "A2"
    assert ops[1].address == "A10:F100"


def test_write_values_snapshots_scalar_at_queue_time() -> None:
    plan = WritePlan()
    x = 5
    plan.write_values("Sheet", "A1", x)
    x = 10  # noqa: F841 - rebinding must not change the queued op

    assert list(plan)[0].values == 5


def test_write_values_snapshots_mutable_values_at_queue_time() -> None:
    plan = WritePlan()
    row = [1, 2, 3]
    plan.write_values("Sheet", "A1", row)

    # In-place mutation after queueing must not leak into the queued op.
    row[0] = 99
    row.append(4)

    assert list(plan)[0].values == [1, 2, 3]


def test_methods_chain_and_clear() -> None:
    plan = (
        WritePlan()
        .write_values("Summary", "B2", "確定")
        .clear_contents("Data", "A1:B2")
    )
    assert len(plan) == 2

    plan.clear()
    assert len(plan) == 0


def test_at_methods_validate_when_queued() -> None:
    plan = WritePlan()
    with pytest.raises(ValueError):
        plan.write_values_at("Data", row=0, column=1, values="x")
    assert len(plan) == 0


def test_apply_dispatches_each_op_then_saves(tmp_path: Path) -> None:
    path = tmp_path / "report.xlsx"
    make_workbook(path)

    plan = WritePlan()
    plan.write_values("Summary", "B2", "更新済み")
    plan.clear_contents("Data", "A2:F1000")
    plan.write_values("Data", "A2", [["Alice", 95], ["Bob", 88]])

    with ExcelWorkbook(path) as book:
        recorder = RecordingWriter()
        book._writer = recorder
        assert book.read_cell("Summary", "A1") == "title"
        book.apply(plan)

    assert recorder.calls == [
        ("write_values", "Summary", "B2", "更新済み", False),
        ("clear_contents", "Data", "A2:F1000"),
        ("write_values", "Data", "A2", [["Alice", 95], ["Bob", 88]], False),
    ]
    assert recorder.saved == 1


def test_apply_is_non_destructive_and_reusable(tmp_path: Path) -> None:
    path = tmp_path / "report.xlsx"
    make_workbook(path)

    plan = WritePlan().write_values("Summary", "B2", "確定")

    with ExcelWorkbook(path) as book:
        recorder = RecordingWriter()
        book._writer = recorder
        book.apply(plan)
        book.apply(plan)

    assert len(plan) == 1
    assert recorder.calls.count(("write_values", "Summary", "B2", "確定", False)) == 2
    assert recorder.saved == 2


def test_apply_with_save_false_skips_save(tmp_path: Path) -> None:
    path = tmp_path / "report.xlsx"
    make_workbook(path)

    plan = WritePlan().write_values("Summary", "B2", "確定")

    with ExcelWorkbook(path) as book:
        recorder = RecordingWriter()
        book._writer = recorder
        book.apply(plan, save=False)

    assert recorder.saved == 0
    assert recorder.calls == [("write_values", "Summary", "B2", "確定", False)]


def make_bordered_table() -> BorderTable:
    return BorderTable(
        workbook=None,  # type: ignore[arg-type]  # only used by BorderTable.save()
        sheet="Report",
        start_row=2,
        start_column=1,
        columns=[["h", "row1"], ["amount", 100]],
        header_rows=1,
        header_columns=1,
    )


def test_add_bordered_table_snapshots_current_state() -> None:
    table = make_bordered_table()
    table.add_row([200], row_headers=["row2"])

    plan = WritePlan()
    plan.add_bordered_table(table)

    # Mutating the table after queueing must not change the queued snapshot.
    table.set_value(row=1, column=2, value="changed")
    table.add_row([300], row_headers=["row3"])

    op = list(plan)[0]
    assert op.sheet == "Report"
    assert op.partial is False
    assert op.start_row == 2
    assert op.start_column == 1
    assert op.columns == ((1, ("h", "row1", "row2")), (2, ("amount", 100, 200)))
    assert op.end_row == 4
    assert op.end_column == 2
    assert op.insertions == (("row", 4),)


def test_apply_dispatches_bordered_table(tmp_path: Path) -> None:
    path = tmp_path / "report.xlsx"
    make_workbook(path)

    table = make_bordered_table()
    table.add_row([200], row_headers=["row2"])
    plan = WritePlan().add_bordered_table(table)

    with ExcelWorkbook(path) as book:
        recorder = RecordingWriter()
        book._writer = recorder
        book.apply(plan)

    assert recorder.calls == [
        (
            "apply_bordered_table",
            "Report",
            False,
            2,
            1,
            1,
            [(1, ["h", "row1", "row2"]), (2, ["amount", 100, 200])],
            4,
            2,
            (("row", 4),),
        )
    ]
    assert recorder.saved == 1


def test_empty_plan_never_touches_writer(tmp_path: Path) -> None:
    path = tmp_path / "report.xlsx"
    make_workbook(path)

    with ExcelWorkbook(path) as book:
        recorder = RecordingWriter()
        book._writer = recorder
        book.apply(WritePlan())

    assert recorder.calls == []
    assert recorder.saved == 0
