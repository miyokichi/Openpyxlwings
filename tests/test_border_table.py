from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import Border, Side

from openpyxlwings import BorderTable, ExcelWorkbook
from openpyxlwings.exceptions import BorderTableShapeError


def make_bordered_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"

    values = [
        ["", "2026", "2027"],
        ["Region", "Sales", "Sales"],
        ["East", 10, 20],
        ["West", 30, 40],
    ]
    thin = Side(style="thin")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    for row_offset, row in enumerate(values, start=2):
        for column_offset, value in enumerate(row, start=2):
            cell = sheet.cell(row=row_offset, column=column_offset)
            cell.value = value
            cell.border = border

    workbook.save(path)


def make_amount_header_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Amount"

    values = [
        ["header1", "header2", "amount", "amount"],
        ["header_col1", "header2_col1", 100, 200],
        ["header_col2", "header2_col2", 300, 400],
        ["header_col3", "header2_col3", 500, 600],
    ]
    thin = Side(style="thin")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for row_offset, row in enumerate(values, start=2):
        for column_offset, value in enumerate(row, start=2):
            cell = sheet.cell(row=row_offset, column=column_offset)
            cell.value = value
            cell.border = border

    workbook.save(path)


def test_get_bordered_table_detects_range_and_headers(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    make_bordered_workbook(path)

    with ExcelWorkbook(path) as workbook:
        table = workbook.get_bordered_table(
            "Report",
            row=3,
            column=3,
            header_rows=2,
            header_columns=1,
        )

    assert table.range == "B2:D5"
    assert table.columns == [
        [None, "Region", "East", "West"],
        ["2026", "Sales", 10, 30],
        ["2027", "Sales", 20, 40],
    ]
    assert table.column_headers == [["2026", "2027"], ["Sales", "Sales"]]
    assert table.row_headers == [["East"], ["West"]]
    assert table.data == [[10, 30], [20, 40]]


def test_get_bordered_table_by_header_detects_variable_amount_columns(tmp_path: Path) -> None:
    path = tmp_path / "amount.xlsx"
    make_amount_header_workbook(path)

    with ExcelWorkbook(path) as workbook:
        table = workbook.get_bordered_table(
            "Amount",
            header_values=["header1", "header2"],
            value_header_contains="amount",
        )

    assert table.range == "B2:E5"
    assert table.header_rows == 1
    assert table.header_columns == 2
    assert table.column_headers == [["amount", "amount"]]
    assert table.row_headers == [
        ["header_col1", "header2_col1"],
        ["header_col2", "header2_col2"],
        ["header_col3", "header2_col3"],
    ]
    assert table.data == [[100, 300, 500], [200, 400, 600]]


def test_get_bordered_table_by_header_skips_decoy_header_text(tmp_path: Path) -> None:
    # The same header sequence appears as plain borderless text above the real
    # table; the decoy must be skipped instead of aborting the search.
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Amount"
    sheet.cell(row=1, column=2).value = "header1"
    sheet.cell(row=1, column=3).value = "header2"

    values = [
        ["header1", "header2", "amount"],
        ["header_col1", "header2_col1", 100],
        ["header_col2", "header2_col2", 300],
    ]
    thin = Side(style="thin")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for row_offset, row in enumerate(values, start=3):
        for column_offset, value in enumerate(row, start=2):
            cell = sheet.cell(row=row_offset, column=column_offset)
            cell.value = value
            cell.border = border
    path = tmp_path / "decoy.xlsx"
    workbook.save(path)

    with ExcelWorkbook(path) as book:
        table = book.get_bordered_table(
            "Amount",
            header_values=["header1", "header2"],
            value_header_contains="amount",
        )

    assert table.range == "B3:D5"


def test_get_bordered_table_by_header_is_case_insensitive(tmp_path: Path) -> None:
    path = tmp_path / "amount.xlsx"
    make_amount_header_workbook(path)

    with ExcelWorkbook(path) as workbook:
        table = workbook.get_bordered_table(
            "Amount",
            header_values=["HEADER1", "HEADER2"],
            value_header_contains="AMOUNT",
        )

    assert table.range == "B2:E5"


def test_get_bordered_table_rejects_bad_argument_combinations(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    make_bordered_workbook(path)

    with ExcelWorkbook(path) as workbook:
        with pytest.raises(BorderTableShapeError, match="either row/column or header_values"):
            workbook.get_bordered_table("Report")
        with pytest.raises(BorderTableShapeError, match="either row/column or header_values"):
            workbook.get_bordered_table("Report", row=3, column=3, header_values=["Region"])
        with pytest.raises(BorderTableShapeError, match="given together"):
            workbook.get_bordered_table("Report", row=3)
        with pytest.raises(BorderTableShapeError, match="requires header_values"):
            workbook.get_bordered_table("Report", row=3, column=3, columns="selected")
        with pytest.raises(BorderTableShapeError, match="requires header_values"):
            workbook.get_bordered_table(
                "Report", row=3, column=3, value_header_contains="Sales"
            )
        with pytest.raises(BorderTableShapeError, match="do not pass it"):
            workbook.get_bordered_table(
                "Report",
                header_values=["Region"],
                value_header_contains="Sales",
                header_columns=1,
            )
        with pytest.raises(BorderTableShapeError, match="value_header_contains is required"):
            workbook.get_bordered_table("Report", header_values=["Region"])
        with pytest.raises(BorderTableShapeError, match='"all" or "selected"'):
            workbook.get_bordered_table("Report", row=3, column=3, columns="everything")


def test_border_table_edit_methods_update_values(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    make_bordered_workbook(path)

    with ExcelWorkbook(path) as workbook:
        table = workbook.get_bordered_table("Report", row=4, column=4, header_rows=2, header_columns=1)

    table.set_value(1, 2, "FY2026")
    table.set_body_value(2, 2, 99)
    table.add_row([50, 60], row_headers=["North"])
    table.add_column([70, 80, 90], column_headers=["2028", "Sales"])
    table.add_header_row(["Area", "Actual", "Actual", "Plan"])
    table.add_header_column(["Group", "Metric", "Metric", "A", "B", "C"])

    assert table.header_rows == 3
    assert table.header_columns == 2
    assert table.columns == [
        [None, "Region", "Area", "East", "West", "North"],
        ["Group", "Metric", "Metric", "A", "B", "C"],
        ["FY2026", "Sales", "Actual", 10, 30, 50],
        ["2027", "Sales", "Actual", 20, 99, 60],
        ["2028", "Sales", "Plan", 70, 80, 90],
    ]


def test_border_table_sets_body_row_by_single_row_header(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    make_bordered_workbook(path)

    with ExcelWorkbook(path) as workbook:
        table = workbook.get_bordered_table("Report", row=4, column=4, header_rows=2, header_columns=1)

    assert table.find_body_row("West") == 2

    table.set_body_row_by_header("East", [11, 22])

    assert table.data == [[11, 30], [22, 40]]


def test_border_table_sets_body_row_by_multi_column_row_header() -> None:
    table = BorderTable(
        workbook=None,  # type: ignore[arg-type]
        sheet="Report",
        start_row=1,
        start_column=1,
        columns=[
            ["Region", "East", "East", "West"],
            ["Segment", "Retail", "Enterprise", "Retail"],
            ["Sales", 10, 20, 30],
            ["Cost", 3, 8, 9],
        ],
        header_rows=1,
        header_columns=2,
    )

    assert table.find_body_row(["East", "Enterprise"]) == 2

    table.set_body_row_by_header(("East", "Enterprise"), [25, 10])

    assert table.data == [[10, 25, 30], [3, 10, 9]]


def test_border_table_rejects_bad_shapes(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    make_bordered_workbook(path)

    with ExcelWorkbook(path) as workbook:
        table = workbook.get_bordered_table("Report", row=4, column=4, header_rows=2, header_columns=1)

    with pytest.raises(BorderTableShapeError):
        table.set_body_value(10, 1, "outside")

    with pytest.raises(BorderTableShapeError):
        table.add_row([1])

    with pytest.raises(BorderTableShapeError):
        table.add_column([1])


def test_border_table_rejects_bad_row_header_updates() -> None:
    table = BorderTable(
        workbook=None,  # type: ignore[arg-type]
        sheet="Report",
        start_row=1,
        start_column=1,
        columns=[
            ["Region", "East", "East"],
            ["Sales", 10, 20],
        ],
        header_rows=1,
        header_columns=1,
    )

    with pytest.raises(BorderTableShapeError, match="matches multiple"):
        table.find_body_row("East")

    with pytest.raises(BorderTableShapeError, match="was not found"):
        table.find_body_row("West")

    with pytest.raises(BorderTableShapeError, match="row values length"):
        table.set_body_row_by_header("East", [1, 2])

    with pytest.raises(BorderTableShapeError, match="row_header length"):
        table.set_body_row_by_header(["East", "Retail"], [1])


def test_get_bordered_table_rejects_missing_internal_border(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    make_bordered_workbook(path)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Broken"
    thin = Side(style="thin")
    full = Border(top=thin, bottom=thin, left=thin, right=thin)
    missing_right = Border(top=thin, bottom=thin, left=thin)
    for row in range(1, 3):
        for column in range(1, 3):
            sheet.cell(row=row, column=column).border = full
    sheet.cell(row=2, column=1).border = missing_right
    sheet.cell(row=2, column=2).border = Border(top=thin, bottom=thin, right=thin)
    workbook.save(path)

    with ExcelWorkbook(path) as reader:
        with pytest.raises(BorderTableShapeError):
            reader.get_bordered_table("Broken", row=1, column=1)


def make_missing_inner_border_workbook(path: Path) -> None:
    """A 3x3 framed table whose middle vertical gridline is absent."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Partial"
    values = [
        ["Region", "Q1", "Q2"],
        ["East", 10, 20],
        ["West", 30, 40],
    ]
    thin = Side(style="thin")
    full = Border(top=thin, bottom=thin, left=thin, right=thin)
    no_right = Border(top=thin, bottom=thin, left=thin)
    no_left = Border(top=thin, bottom=thin, right=thin)
    for row_offset, row in enumerate(values, start=2):       # rows 2..4
        for column_offset, value in enumerate(row, start=2):  # cols 2..4 (table cols 1..3)
            cell = sheet.cell(row=row_offset, column=column_offset)
            cell.value = value
            if column_offset == 2:        # left table column: drop its right edge
                cell.border = no_right
            elif column_offset == 3:      # middle column: drop its left edge
                cell.border = no_left
            else:                         # right table column: full
                cell.border = full
    workbook.save(path)


def test_require_inner_borders_false_reads_table_with_missing_inner_border(
    tmp_path: Path,
) -> None:
    path = tmp_path / "book.xlsx"
    make_missing_inner_border_workbook(path)

    with ExcelWorkbook(path) as reader:
        table = reader.get_bordered_table(
            "Partial",
            row=3,
            column=3,
            header_rows=1,
            header_columns=1,
            require_inner_borders=False,
        )

    assert table.range == "B2:D4"
    assert table.columns == [
        ["Region", "East", "West"],
        ["Q1", 10, 30],
        ["Q2", 20, 40],
    ]


def test_missing_inner_border_still_rejected_by_default(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    make_missing_inner_border_workbook(path)

    with ExcelWorkbook(path) as reader:
        with pytest.raises(BorderTableShapeError):
            reader.get_bordered_table("Partial", row=3, column=3, header_rows=1, header_columns=1)


def test_require_inner_borders_false_reads_with_borderless_inner_cell(
    tmp_path: Path,
) -> None:
    path = tmp_path / "book.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Hole"
    values = [
        ["Region", "Q1", "Q2"],
        ["East", 10, 20],
        ["West", 30, 40],
    ]
    thin = Side(style="thin")
    full = Border(top=thin, bottom=thin, left=thin, right=thin)
    for row_offset, row in enumerate(values, start=2):
        for column_offset, value in enumerate(row, start=2):
            cell = sheet.cell(row=row_offset, column=column_offset)
            cell.value = value
            cell.border = full
    # Center body cell loses every border.
    sheet.cell(row=3, column=3).border = Border()
    workbook.save(path)

    with ExcelWorkbook(path) as reader:
        table = reader.get_bordered_table(
            "Hole",
            row=2,
            column=2,
            header_rows=1,
            header_columns=1,
            require_inner_borders=False,
        )

    assert table.range == "B2:D4"
    assert table.columns[1][1] == 10


def test_border_table_save_delegates_to_workbook() -> None:
    class FakeWorkbook:
        def __init__(self) -> None:
            self.saved_table = None

        def _save_bordered_table(self, table: BorderTable) -> None:
            self.saved_table = table

    workbook = FakeWorkbook()
    table = BorderTable(
        workbook=workbook,  # type: ignore[arg-type]
        sheet="Report",
        start_row=1,
        start_column=1,
        columns=[["Name", "Alice"], ["Score", 10]],
    )
    table.add_row(["Bob", 20])

    table.save()

    assert workbook.saved_table is table
    assert table._insertions == []


def test_repeated_bordered_table_reads_reuse_single_file_open(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "book.xlsx"
    make_bordered_workbook(path)

    import openpyxlwings.workbook as workbook_module

    styled_opens = 0
    real_load_workbook = workbook_module.load_workbook

    def counting_load_workbook(*args, **kwargs):
        nonlocal styled_opens
        if kwargs.get("read_only") is False:
            styled_opens += 1
        return real_load_workbook(*args, **kwargs)

    monkeypatch.setattr(workbook_module, "load_workbook", counting_load_workbook)

    with ExcelWorkbook(path) as workbook:
        first = workbook.get_bordered_table("Report", row=3, column=3, header_rows=2, header_columns=1)
        second = workbook.get_bordered_table("Report", row=3, column=3, header_rows=2, header_columns=1)

    assert first.range == second.range == "B2:D5"
    assert styled_opens == 1
