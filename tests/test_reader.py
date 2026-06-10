from pathlib import Path

from openpyxl import Workbook

from openpyxlwings import (
    ExcelWorkbook,
    read_cell_at,
    read_range,
    read_range_at,
    read_sheet,
    sheet_names,
)


def make_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = "name"
    sheet["B1"] = "score"
    sheet["A2"] = "alice"
    sheet["B2"] = 10
    sheet["D5"] = None
    workbook.create_sheet("Other")
    workbook.save(path)


def test_sheet_names(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    make_workbook(path)

    assert sheet_names(path) == ["Data", "Other"]


def test_read_range(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    make_workbook(path)

    assert read_range(path, "Data", "A1:B2") == [["name", "score"], ["alice", 10]]


def test_read_sheet_trims_empty_edges(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    make_workbook(path)

    assert read_sheet(path, "Data") == [["name", "score"], ["alice", 10]]


def test_reader_context_manager(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    make_workbook(path)

    with ExcelWorkbook(path) as workbook:
        assert workbook.read_cell("Data", "A2") == "alice"


def test_read_cell_at_uses_row_and_column_numbers(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    make_workbook(path)

    with ExcelWorkbook(path) as workbook:
        assert workbook.read_cell_at("Data", 2, 2) == 10


def test_read_range_at_uses_row_and_column_numbers(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    make_workbook(path)

    with ExcelWorkbook(path) as workbook:
        assert workbook.read_range_at("Data", 1, 1, 2, 2) == [
            ["name", "score"],
            ["alice", 10],
        ]


def test_numeric_helper_functions(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    make_workbook(path)

    assert read_cell_at(path, "Data", 2, 1) == "alice"
    assert read_range_at(path, "Data", 1, 1, 2, 2) == [
        ["name", "score"],
        ["alice", 10],
    ]


def test_row_and_column_numbers_are_one_based(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    make_workbook(path)

    with ExcelWorkbook(path) as workbook:
        try:
            workbook.read_cell_at("Data", 0, 1)
        except ValueError as exc:
            assert "row must be 1 or greater" in str(exc)
        else:
            raise AssertionError("Expected ValueError")
