"""Tests for partial (column-selected) bordered tables."""

from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import Border, Side

from openpyxlwings import BorderTable, ExcelWorkbook, WritePlan
from openpyxlwings.exceptions import BorderTableNotFoundError, BorderTableShapeError
from openpyxlwings.plan import _BorderedTableOp
from openpyxlwings.workbook import _XlwingsWriteSession

THIN = Side(style="thin")
FULL = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def _write_grid(path: Path, title: str, values) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    for row_offset, row in enumerate(values, start=2):
        for column_offset, value in enumerate(row, start=2):
            cell = sheet.cell(row=row_offset, column=column_offset)
            cell.value = value
            cell.border = FULL
    workbook.save(path)


def make_amount_workbook(path: Path) -> None:
    _write_grid(
        path,
        "Amount",
        [
            ["header1", "header2", "amount", "amount"],
            ["col1a", "col1b", 100, 200],
            ["col2a", "col2b", 300, 400],
            ["col3a", "col3b", 500, 600],
        ],
    )


def make_partial_table(workbook=None) -> BorderTable:
    return BorderTable(
        workbook=workbook,
        sheet="Amount",
        start_row=2,
        start_column=2,
        columns=[["key", "a", "b"], ["amount", 100, 200]],
        header_rows=1,
        header_columns=1,
        source_columns=[2, 3],
        partial=True,
        detected_end_row=4,
        detected_end_column=3,
    )


def test_selects_only_requested_header_column(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    make_amount_workbook(path)

    with ExcelWorkbook(path) as workbook:
        table = workbook.get_bordered_table("Amount", header_values=["header1"], columns="selected")

    assert table.partial is True
    assert [column[0] for column in table.columns] == ["header1"]
    assert table.row_headers == [["col1a"], ["col2a"], ["col3a"]]
    assert table.data == []  # every held column is a row-header column
    assert table.row_count == 4  # header row + 3 body rows


def test_value_header_contains_selects_every_amount_column(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    make_amount_workbook(path)

    with ExcelWorkbook(path) as workbook:
        table = workbook.get_bordered_table(
            "Amount",
            header_values=["header1"],
            value_header_contains="amount",
            columns="selected",
        )

    # header2 is not requested, so it is excluded; both amount columns appear.
    assert table.column_headers == [["amount", "amount"]]
    assert table.row_headers == [["col1a"], ["col2a"], ["col3a"]]
    assert table.data == [[100, 300, 500], [200, 400, 600]]
    # Each held column remembers its source Excel column (B, C, D).
    assert table.source_columns == [2, 4, 5]


def test_continuation_includes_bordered_empty_cell(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    _write_grid(
        path,
        "Gap",
        [
            ["key", "amount"],
            ["a", 100],
            ["b", None],
            ["c", 300],
        ],
    )

    with ExcelWorkbook(path) as workbook:
        table = workbook.get_bordered_table(
            "Gap",
            header_values=["key"],
            value_header_contains="amount",
            columns="selected",
        )

    assert table.row_headers == [["a"], ["b"], ["c"]]
    assert table.data == [[100, None, 300]]


def test_header_on_second_table_row_keeps_title_row(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    _write_grid(
        path,
        "Titled",
        [
            ["Title", None, None],
            ["key", "amount", "note"],
            ["a", 100, "x"],
            ["b", 200, "y"],
        ],
    )

    with ExcelWorkbook(path) as workbook:
        table = workbook.get_bordered_table(
            "Titled",
            header_values=["key"],
            value_header_contains="amount",
            columns="selected",
            header_rows=2,
        )

    assert table.header_rows == 2
    assert table.columns == [["Title", "key", "a", "b"], [None, "amount", 100, 200]]
    assert table.row_headers == [["a"], ["b"]]
    assert table.data == [[100, 200]]


def test_ragged_columns_are_padded_with_none() -> None:
    # Columns read with different lengths are squared up into a rectangle.
    table = BorderTable(
        workbook=None,
        sheet="S",
        start_row=2,
        start_column=2,
        columns=[["key", "a", "b", "c"], ["amount", 100]],
        header_rows=1,
        header_columns=1,
        source_columns=[2, 3],
        partial=True,
        detected_end_row=5,
        detected_end_column=3,
    )

    assert table.row_count == 4
    assert table.row_headers == [["a"], ["b"], ["c"]]
    assert table.data == [[100, None, None]]


def test_partial_requires_detected_bounds() -> None:
    with pytest.raises(BorderTableShapeError, match="detected_end_row"):
        BorderTable(
            workbook=None,
            sheet="S",
            start_row=2,
            start_column=2,
            columns=[["key", "a"]],
            source_columns=[2],
            partial=True,
        )


def test_decoy_first_header_on_same_row_is_skipped(tmp_path: Path) -> None:
    # A lone cell with the same text sits in the same row, separated from the
    # real table by an empty column; the search must move on to the next
    # candidate cell instead of stopping at the decoy.
    path = tmp_path / "book.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Amount"
    sheet.cell(row=2, column=2).value = "header1"  # decoy, no table below
    values = [
        ["header1", "amount"],
        ["col1a", 100],
        ["col2a", 300],
        ["col3a", 500],
    ]
    for row_offset, row in enumerate(values, start=2):
        for column_offset, value in enumerate(row, start=4):  # columns D..E
            cell = sheet.cell(row=row_offset, column=column_offset)
            cell.value = value
            cell.border = FULL
    workbook.save(path)

    with ExcelWorkbook(path) as workbook:
        table = workbook.get_bordered_table("Amount", header_values=["header1"], columns="selected")

    assert table.source_columns == [4]
    assert [column[0] for column in table.columns] == ["header1"]
    assert table.row_headers == [["col1a"], ["col2a"], ["col3a"]]


def make_metrics_workbook(path: Path) -> None:
    _write_grid(
        path,
        "Metrics",
        [
            ["metric", "prodA", "prodB", "prodC", "prodD"],
            ["sales", 120, 80, 200, 50],
            ["rate", 0.9, 0.6, 1.1, 0.4],
            ["flag", "OK", "NG", "ok", None],
        ],
    )


def load_metrics_table(path: Path) -> BorderTable:
    make_metrics_workbook(path)
    with ExcelWorkbook(path) as workbook:
        return workbook.get_bordered_table("Metrics", row=2, column=2, header_columns=1)


def test_select_columns_by_row_with_callable_condition(tmp_path: Path) -> None:
    table = load_metrics_table(tmp_path / "book.xlsx")

    subset = table.select_columns_by_row("rate", lambda v: v is not None and v >= 0.8)

    assert subset.partial is True
    assert subset.source_columns == [2, 3, 5]  # metric + prodA + prodC
    assert subset.column_headers == [["prodA", "prodC"]]
    assert subset.row_headers == [["sales"], ["rate"], ["flag"]]
    assert subset.data == [[120, 0.9, "OK"], [200, 1.1, "ok"]]
    assert subset.detected_end_row == table.end_row
    assert subset.detected_end_column == table.end_column


def test_select_columns_by_row_with_plain_value_condition(tmp_path: Path) -> None:
    table = load_metrics_table(tmp_path / "book.xlsx")

    # Plain values match like headers: stripped and case-insensitive.
    subset = table.select_columns_by_row("flag", "ok")

    assert subset.source_columns == [2, 3, 5]
    assert subset.column_headers == [["prodA", "prodC"]]

    strict = table.select_columns_by_row("flag", "ok", match_case=True)
    assert strict.source_columns == [2, 5]


def test_select_columns_by_row_from_partial_table(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    make_metrics_workbook(path)
    with ExcelWorkbook(path) as workbook:
        table = workbook.get_bordered_table(
            "Metrics",
            header_values=["metric"],
            value_header_contains="prod",
            columns="selected",
        )

    subset = table.select_columns_by_row("rate", lambda v: v is not None and v >= 0.8)

    assert subset.source_columns == [2, 3, 5]
    assert subset.data == [[120, 0.9, "OK"], [200, 1.1, "ok"]]


def test_select_columns_by_row_edits_do_not_affect_parent(tmp_path: Path) -> None:
    table = load_metrics_table(tmp_path / "book.xlsx")

    subset = table.select_columns_by_row("rate", lambda v: v is not None and v >= 0.8)
    subset.set_body_value(1, 1, 999)

    assert subset.data[0][0] == 999
    assert table.data[0][0] == 120


def test_select_columns_by_row_rejects_no_match(tmp_path: Path) -> None:
    table = load_metrics_table(tmp_path / "book.xlsx")

    with pytest.raises(BorderTableShapeError, match="no body columns matched"):
        table.select_columns_by_row("rate", lambda value: False)

    with pytest.raises(BorderTableShapeError, match="row_header was not found"):
        table.select_columns_by_row("nope", lambda value: True)


def test_select_columns_by_row_save_writes_only_selected_columns(tmp_path: Path) -> None:
    class FakeWorkbook:
        def __init__(self) -> None:
            self.saved_table = None

        def _save_bordered_table(self, table: BorderTable, path=None) -> None:
            self.saved_table = table

    table = load_metrics_table(tmp_path / "book.xlsx")
    fake = FakeWorkbook()
    table.workbook = fake

    subset = table.select_columns_by_row("rate", lambda v: v is not None and v >= 0.8)
    subset.set_body_row_by_header("flag", ["A", "B"])
    subset.save()

    assert fake.saved_table is subset
    assert fake.saved_table.partial is True
    assert fake.saved_table.source_columns == [2, 3, 5]
    assert subset.data == [[120, 0.9, "A"], [200, 1.1, "B"]]


# --- select_columns_by_column_header ---------------------------------------


def test_select_columns_by_column_header_callable(tmp_path: Path) -> None:
    table = load_metrics_table(tmp_path / "book.xlsx")

    subset = table.select_columns_by_column_header(lambda h: h in ("prodA", "prodC"))

    assert subset.partial_axis == "column"
    assert subset.column_headers == [["prodA", "prodC"]]
    assert subset.source_columns == [2, 3, 5]  # metric + prodA + prodC
    assert subset.row_headers == [["sales"], ["rate"], ["flag"]]
    # data is column-major: one list per kept body column (prodA, prodC).
    assert subset.data == [[120, 0.9, "OK"], [200, 1.1, "ok"]]


def test_select_columns_by_column_header_plain_value(tmp_path: Path) -> None:
    table = load_metrics_table(tmp_path / "book.xlsx")

    subset = table.select_columns_by_column_header("PRODB")  # case-insensitive
    assert subset.source_columns == [2, 4]
    assert subset.column_headers == [["prodB"]]

    strict = table.select_columns_by_column_header("prodB", match_case=True)
    assert strict.source_columns == [2, 4]


def test_select_columns_by_column_header_multi_header_rows() -> None:
    # Two column-header rows: the condition receives a tuple per column.
    table = BorderTable(
        workbook=None,
        sheet="S",
        start_row=1,
        start_column=1,
        columns=[
            ["", "", "East", "West"],  # corner (2 rows) + row headers East/West
            ["2026", "Q1", 10, 40],  # column header (2026, Q1), body 10/40
            ["2026", "Q2", 20, 50],  # (2026, Q2)
            ["2027", "Q1", 30, 60],  # (2027, Q1)
        ],
        header_rows=2,
        header_columns=1,
    )

    subset = table.select_columns_by_column_header(lambda h: h[0] == "2026")

    assert subset.source_columns == [1, 2, 3]  # row-header col + the two 2026 columns
    assert subset.data == [[10, 40], [20, 50]]


def test_select_columns_by_column_header_rejects_no_match(tmp_path: Path) -> None:
    table = load_metrics_table(tmp_path / "book.xlsx")

    with pytest.raises(BorderTableShapeError, match="no body columns matched"):
        table.select_columns_by_column_header(lambda h: False)


def test_select_columns_by_column_header_rejects_row_partial(tmp_path: Path) -> None:
    table = load_metrics_table(tmp_path / "book.xlsx")
    row_subset = table.select_rows_by_row_header(lambda h: h in ("sales", "flag"))

    with pytest.raises(BorderTableShapeError, match="cannot column-filter"):
        row_subset.select_columns_by_column_header(lambda h: True)


# --- select_rows_by_row_header ---------------------------------------------


def test_select_rows_by_row_header_callable(tmp_path: Path) -> None:
    table = load_metrics_table(tmp_path / "book.xlsx")

    subset = table.select_rows_by_row_header(lambda h: h in ("sales", "flag"))

    assert subset.partial_axis == "row"
    assert subset.row_headers == [["sales"], ["flag"]]
    assert subset.source_rows == [2, 3, 5]  # header row 2, sales row 3, flag row 5
    assert subset.source_columns == [2, 3, 4, 5, 6]  # every column kept
    assert subset.column_headers == [["prodA", "prodB", "prodC", "prodD"]]
    assert subset.data == [[120, "OK"], [80, "NG"], [200, "ok"], [50, None]]


def test_select_rows_by_row_header_plain_value(tmp_path: Path) -> None:
    table = load_metrics_table(tmp_path / "book.xlsx")

    subset = table.select_rows_by_row_header("RATE")  # case-insensitive
    assert subset.source_rows == [2, 4]
    assert subset.row_headers == [["rate"]]


def test_select_rows_by_row_header_multi_column_headers() -> None:
    table = BorderTable(
        workbook=None,
        sheet="S",
        start_row=1,
        start_column=1,
        columns=[
            ["region", "East", "East", "West"],
            ["segment", "Retail", "Corp", "Retail"],
            ["sales", 10, 20, 30],
        ],
        header_rows=1,
        header_columns=2,
    )

    subset = table.select_rows_by_row_header(lambda h: h[0] == "East")

    assert subset.source_rows == [1, 2, 3]
    assert subset.row_headers == [["East", "Retail"], ["East", "Corp"]]
    assert subset.data == [[10, 20]]


def test_select_rows_by_row_header_rejects_no_match(tmp_path: Path) -> None:
    table = load_metrics_table(tmp_path / "book.xlsx")

    with pytest.raises(BorderTableShapeError, match="no body rows matched"):
        table.select_rows_by_row_header(lambda h: False)


def test_select_rows_by_row_header_requires_row_headers(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    _write_grid(path, "NoRowHeaders", [["a", "b"], [1, 2], [3, 4]])
    with ExcelWorkbook(path) as workbook:
        table = workbook.get_bordered_table("NoRowHeaders", row=2, column=2, header_columns=0)

    with pytest.raises(BorderTableShapeError, match="no row headers"):
        table.select_rows_by_row_header(lambda h: True)


def test_select_rows_by_row_header_rejects_column_partial(tmp_path: Path) -> None:
    table = load_metrics_table(tmp_path / "book.xlsx")
    col_subset = table.select_columns_by_column_header(lambda h: h in ("prodA", "prodC"))

    with pytest.raises(BorderTableShapeError, match="cannot row-filter"):
        col_subset.select_rows_by_row_header(lambda h: True)


def test_select_rows_by_row_header_edits_do_not_affect_parent(tmp_path: Path) -> None:
    table = load_metrics_table(tmp_path / "book.xlsx")

    subset = table.select_rows_by_row_header(lambda h: h == "sales")
    subset.set_body_value(1, 1, 999)

    assert subset.data[0][0] == 999
    assert table.data[0][0] == 120


def test_select_rows_by_row_header_add_row_and_column(tmp_path: Path) -> None:
    table = load_metrics_table(tmp_path / "book.xlsx")

    subset = table.select_rows_by_row_header(lambda h: h in ("sales", "flag"))
    subset.add_row([1, 2, 3, 4], row_headers=["cost"])
    subset.add_column([9, 8, 9], column_headers=["extra"])

    assert subset.added_rows == 1
    assert subset.added_columns == 1
    assert subset.source_rows == [2, 3, 5, None]
    assert subset.source_columns[-1] is None
    assert subset.end_row == 6  # detected bottom (5) + 1 appended row
    assert subset.end_column == 7  # detected right (6) + 1 appended column
    assert subset.row_headers == [["sales"], ["flag"], ["cost"]]
    assert [column[-1] for column in subset.data] == [1, 2, 3, 4, 9]


def test_select_rows_by_row_header_save_writes_kept_rows(tmp_path: Path) -> None:
    # A row-partial save must write each kept row to its original Excel row,
    # append added rows at the bottom and added columns at the right, and leave
    # unmatched rows untouched.
    class RecordingWriter(_XlwingsWriteSession):
        def __init__(self) -> None:
            self.ops: list[tuple] = []

        def insert_row(self, sheet, row) -> None:
            self.ops.append(("insert_row", row))

        def insert_column(self, sheet, column) -> None:
            self.ops.append(("insert_column", column))

        def write_values_at(self, sheet, row, column, values, *, expand=False) -> None:
            self.ops.append(("write", row, column, values))

        def apply_table_borders(self, sheet, sr, sc, er, ec) -> None:
            self.ops.append(("borders", sr, sc, er, ec))

        def save(self, path=None) -> None:
            self.ops.append(("save", path))

    table = load_metrics_table(tmp_path / "book.xlsx")
    subset = table.select_rows_by_row_header(lambda h: h in ("sales", "flag"))
    subset.add_row([1, 2, 3, 4], row_headers=["cost"])
    subset.add_column([9, 8, 9], column_headers=["extra"])

    writer = RecordingWriter()
    writer.save_bordered_table(subset, None)

    assert writer.ops == [
        ("insert_row", 6),  # added "cost" row at the bottom edge
        ("insert_column", 7),  # added "extra" column at the right edge
        ("write", 2, 2, [["metric", "prodA", "prodB", "prodC", "prodD", "extra"]]),
        ("write", 3, 2, [["sales", 120, 80, 200, 50, 9]]),  # sales at its row
        ("write", 5, 2, [["flag", "OK", "NG", "ok", None, 8]]),  # flag; rate (row 4) untouched
        ("write", 6, 2, [["cost", 1, 2, 3, 4, 9]]),  # appended row
        ("borders", 2, 2, 6, 7),
        ("save", None),
    ]


def test_select_rows_by_row_header_save_rebaselines(tmp_path: Path) -> None:
    class FakeWorkbook:
        def _save_bordered_table(self, table: BorderTable, path=None) -> None:
            pass

    table = load_metrics_table(tmp_path / "book.xlsx")
    table.workbook = FakeWorkbook()
    subset = table.select_rows_by_row_header(lambda h: h in ("sales", "flag"))
    subset.add_row([1, 2, 3, 4], row_headers=["cost"])
    subset.add_column([9, 8, 9], column_headers=["extra"])
    subset.save()

    assert subset.added_rows == 0
    assert subset.added_columns == 0
    assert subset.source_rows == [2, 3, 5, 6]  # appended row now lives at row 6
    assert subset.source_columns == [2, 3, 4, 5, 6, 7]  # appended column at col 7
    assert subset.detected_end_row == 6
    assert subset.detected_end_column == 7


def test_write_plan_snapshots_row_partial_table(tmp_path: Path) -> None:
    table = load_metrics_table(tmp_path / "book.xlsx")
    subset = table.select_rows_by_row_header(lambda h: h in ("sales", "flag"))

    plan = WritePlan()
    plan.add_bordered_table(subset)

    subset.add_row([1, 2, 3, 4], row_headers=["cost"])  # must not change snapshot

    op = next(iter(plan))
    assert isinstance(op, _BorderedTableOp)
    assert op.partial_axis == "row"
    assert op.source_rows == (2, 3, 5)


def test_selected_columns_read_borderless_values(tmp_path: Path) -> None:
    # Column selection no longer needs any borders; the body of each column
    # is read while values continue.
    path = tmp_path / "plain.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Plain"
    values = [
        ["header1", "header2", "amount"],
        ["col1a", "col1b", 100],
        ["col2a", "col2b", 300],
    ]
    for row_offset, row in enumerate(values, start=2):
        for column_offset, value in enumerate(row, start=2):
            sheet.cell(row=row_offset, column=column_offset).value = value
    workbook.save(path)

    with ExcelWorkbook(path) as book:
        table = book.get_bordered_table(
            "Plain",
            header_values=["header1"],
            value_header_contains="amount",
            columns="selected",
        )

    assert table.source_columns == [2, 4]
    assert table.row_headers == [["col1a"], ["col2a"]]
    assert table.data == [[100, 300]]


def test_missing_header_raises(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    make_amount_workbook(path)

    with ExcelWorkbook(path) as workbook:
        with pytest.raises(BorderTableNotFoundError):
            workbook.get_bordered_table("Amount", header_values=["nope"], columns="selected")


def test_add_row_and_add_column_update_partial_table(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    make_amount_workbook(path)

    with ExcelWorkbook(path) as workbook:
        table = workbook.get_bordered_table(
            "Amount",
            header_values=["header1"],
            value_header_contains="amount",
            columns="selected",
        )

    table.add_row([700, 800], row_headers=["col4a"])
    assert table.row_count == 5
    assert table.added_rows == 1
    assert [column[-1] for column in table.columns] == ["col4a", 700, 800]
    assert table.end_row == 6  # detected bottom (row 5) plus one appended row

    table.add_column([1, 2, 3, 4], column_headers=["ratio"])
    assert table.column_headers == [["amount", "amount", "ratio"]]
    assert table.source_columns[-1] is None
    assert table.added_columns == 1
    assert [column[0] for column in table.data] == [100, 200, 1]


def test_find_body_row_and_set_body_row_by_header(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    make_amount_workbook(path)

    with ExcelWorkbook(path) as workbook:
        table = workbook.get_bordered_table(
            "Amount",
            header_values=["header1"],
            value_header_contains="amount",
            columns="selected",
        )

    assert table.row_headers == [["col1a"], ["col2a"], ["col3a"]]
    assert table.find_body_row("col2a") == 2

    table.set_body_row_by_header("col1a", [111, 222])

    assert table.data == [[111, 300, 500], [222, 400, 600]]


def test_find_body_row_with_multi_column_row_header() -> None:
    table = BorderTable(
        workbook=None,
        sheet="S",
        start_row=1,
        start_column=1,
        columns=[
            ["header1", "col1a", "col2a", "col3a"],
            ["header2", "col1b", "col2b", "col3b"],
            ["amount", 100, 300, 500],
        ],
        header_rows=1,
        header_columns=2,
        source_columns=[1, 2, 3],
        partial=True,
        detected_end_row=4,
        detected_end_column=4,
    )

    assert table.find_body_row(["col2a", "col2b"]) == 2

    table.set_body_row_by_header(("col1a", "col1b"), [999])

    assert table.data == [[999, 300, 500]]


def test_row_header_lookup_rejects_bad_row_header(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    make_amount_workbook(path)

    with ExcelWorkbook(path) as workbook:
        table = workbook.get_bordered_table(
            "Amount",
            header_values=["header1"],
            value_header_contains="amount",
            columns="selected",
        )

    with pytest.raises(BorderTableShapeError, match="was not found"):
        table.find_body_row("nope")

    with pytest.raises(BorderTableShapeError, match="row_header length"):
        table.find_body_row(["col1a", "extra"])

    with pytest.raises(BorderTableShapeError, match="row values length"):
        table.set_body_row_by_header("col1a", [1])


def test_row_headers_empty_without_header_columns() -> None:
    table = BorderTable(
        workbook=None,
        sheet="S",
        start_row=2,
        start_column=2,
        columns=[["key", "a", "b", "c"], ["amount", 100, 200, 300]],
        header_rows=1,
        header_columns=0,
        source_columns=[2, 3],
        partial=True,
        detected_end_row=5,
        detected_end_column=3,
    )

    assert table.row_headers == []
    with pytest.raises(BorderTableShapeError, match="no row headers"):
        table.find_body_row("a")


def test_add_row_rejects_wrong_width(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    make_amount_workbook(path)

    with ExcelWorkbook(path) as workbook:
        table = workbook.get_bordered_table("Amount", header_values=["header1"], columns="selected")

    with pytest.raises(BorderTableShapeError):
        table.add_row(["too", "many"])


def test_partial_table_rejects_header_additions() -> None:
    table = make_partial_table()

    with pytest.raises(BorderTableShapeError, match="header rows"):
        table.add_header_row(["x", "y"])
    with pytest.raises(BorderTableShapeError, match="header columns"):
        table.add_header_column(["x", "y", "z"])


def test_save_delegates_to_workbook_and_rebaselines() -> None:
    class FakeWorkbook:
        def __init__(self) -> None:
            self.saved_table = None

        def _save_bordered_table(self, table: BorderTable, path=None) -> None:
            self.saved_table = table

    workbook = FakeWorkbook()
    table = make_partial_table(workbook)
    table.add_row([300], row_headers=["c"])
    table.add_column([9, 9, 9], column_headers=["extra"])
    table.save()

    assert workbook.saved_table is table
    assert table.added_rows == 0  # save() rebaselines the original row count
    assert table.added_columns == 0  # the appended column now has a source
    assert table.detected_end_row == 5
    assert table.source_columns == [2, 3, 4]
    assert table.detected_end_column == 4
    assert table._insertions == []


def test_writer_save_bordered_table_forwards_path_to_save() -> None:
    # The writer applies the edit, then saves; a path must reach save() so the
    # workbook is written to a separate file (Save As).
    class RecordingWriter(_XlwingsWriteSession):
        def __init__(self) -> None:
            self.applied = False
            self.save_path = "unset"

        def apply_bordered_table(self, *args, **kwargs) -> None:
            self.applied = True

        def save(self, path=None) -> None:
            self.save_path = path

    writer = RecordingWriter()
    writer.save_bordered_table(make_partial_table(), "out.xlsx")

    assert writer.applied is True
    assert writer.save_path == "out.xlsx"

    writer.save_bordered_table(make_partial_table())
    assert writer.save_path is None


def test_write_plan_snapshots_partial_table() -> None:
    table = make_partial_table()

    plan = WritePlan()
    plan.add_bordered_table(table)

    # Mutating the table afterwards must not change the queued operation.
    table.add_row([300], row_headers=["c"])
    table.add_column([9, 9, 9], column_headers=["extra"])

    assert len(plan) == 1
    op = next(iter(plan))
    assert isinstance(op, _BorderedTableOp)
    assert op.partial is True
    assert op.insertions == ()
    assert op.columns == (
        (2, ("key", "a", "b")),
        (3, ("amount", 100, 200)),
    )
