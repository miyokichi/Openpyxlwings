"""End-to-end WritePlan tests that drive a real Excel install via xlwings.

Run these on a machine with Microsoft Excel installed:

    uv run pytest tests/test_plan_excel.py -m excel

They are auto-skipped when xlwings cannot start Excel (e.g. CI), so the normal
``uv run pytest`` stays green everywhere while still exercising the real write
path where Excel is present.
"""

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side

from openpyxlwings import ExcelWorkbook, WritePlan

xw = pytest.importorskip("xlwings")

pytestmark = pytest.mark.excel


@pytest.fixture(scope="module")
def excel_app():
    """One shared, hidden Excel instance for the module."""

    try:
        app = xw.App(visible=False, add_book=False)
    except Exception as exc:  # pragma: no cover - depends on local Excel.
        pytest.skip(f"Excel is not available: {exc}")
    app.display_alerts = False
    app.screen_updating = False
    try:
        yield app
    finally:
        app.quit()


def make_workbook(path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary["A1"] = "title"
    data = workbook.create_sheet("Data")
    # Pre-existing content that clear_contents must remove.
    data["A2"] = "stale"
    data["B2"] = 999
    workbook.save(path)


def make_bordered_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"
    values = [
        ["Region", "Sales"],
        ["East", 10],
        ["West", 30],
    ]
    thin = Side(style="thin")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for row_offset, row in enumerate(values, start=2):
        for column_offset, value in enumerate(row, start=2):
            cell = sheet.cell(row=row_offset, column=column_offset)
            cell.value = value
            cell.border = border
    workbook.save(path)


def read_back(path: Path) -> Workbook:
    return load_workbook(path, data_only=True)


def test_apply_writes_and_saves_to_disk(tmp_path: Path, excel_app) -> None:
    path = tmp_path / "report.xlsx"
    make_workbook(path)

    plan = WritePlan()
    plan.write_values("Summary", "B2", "更新済み")
    plan.clear_contents("Data", "A2:F1000")
    plan.write_values("Data", "A2", [["Alice", 95], ["Bob", 88]])

    with ExcelWorkbook(path, app=excel_app) as book:
        assert book.read_cell("Summary", "A1") == "title"
        book.apply(plan)

    result = read_back(path)
    assert result["Summary"]["B2"].value == "更新済み"
    assert result["Data"]["A2"].value == "Alice"
    assert result["Data"]["B2"].value == 95
    assert result["Data"]["A3"].value == "Bob"


def test_apply_at_methods_land_on_disk(tmp_path: Path, excel_app) -> None:
    path = tmp_path / "report.xlsx"
    make_workbook(path)

    plan = (
        WritePlan()
        .write_values_at("Data", row=5, column=2, values=[[1, 2], [3, 4]])
        .clear_contents_at("Data", start_row=2, start_column=1, end_row=2, end_column=2)
    )

    with ExcelWorkbook(path, app=excel_app) as book:
        book.apply(plan)

    result = read_back(path)
    assert result["Data"]["B5"].value == 1
    assert result["Data"]["C5"].value == 2
    assert result["Data"]["B6"].value == 3
    assert result["Data"]["C6"].value == 4
    # The cleared cells are gone.
    assert result["Data"]["A2"].value is None
    assert result["Data"]["B2"].value is None


def test_apply_save_false_is_still_saved_on_exit(tmp_path: Path, excel_app) -> None:
    path = tmp_path / "report.xlsx"
    make_workbook(path)

    plan = WritePlan().write_values("Summary", "B2", "確定")

    with ExcelWorkbook(path, app=excel_app) as book:
        book.apply(plan, save=False)

    # The context manager saves on a clean exit even though apply() did not.
    assert read_back(path)["Summary"]["B2"].value == "確定"


def test_bordered_table_edit_via_plan_lands_on_disk(tmp_path: Path, excel_app) -> None:
    path = tmp_path / "report.xlsx"
    make_bordered_workbook(path)

    plan = WritePlan()

    # Detection is openpyxl-only (no Excel), so it can happen before the write
    # session; the edits accumulate in memory until apply().
    with ExcelWorkbook(path) as book:
        table = book.get_bordered_table("Report", row=2, column=2, header_rows=1, header_columns=1)
        table.set_value(row=2, column=2, value=99)        # East sales 10 -> 99
        table.add_row([50], row_headers=["South"])
        plan.add_bordered_table(table)

    with ExcelWorkbook(path, app=excel_app) as book:
        book.apply(plan)

    result = read_back(path)["Report"]
    assert result["C3"].value == 99           # edited East sales
    assert result["B5"].value == "South"      # appended row header
    assert result["C5"].value == 50           # appended row value
    assert result["C4"].value == 30           # West sales pushed down by the insert


def test_plan_reused_across_workbooks(tmp_path: Path, excel_app) -> None:
    plan = WritePlan().write_values("Summary", "B2", "共通")

    paths = [tmp_path / "a.xlsx", tmp_path / "b.xlsx"]
    for path in paths:
        make_workbook(path)
        with ExcelWorkbook(path, app=excel_app) as book:
            book.apply(plan)

    assert len(plan) == 1  # not consumed
    for path in paths:
        assert read_back(path)["Summary"]["B2"].value == "共通"
