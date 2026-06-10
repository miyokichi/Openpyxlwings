"""Create a sample workbook for manual openpyxlwings testing."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "samples" / "openpyxlwings_sample.xlsx"


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)

    create_quick_read_write_sheet(workbook)
    create_bordered_table_sheet(workbook)
    create_broken_table_sheet(workbook)

    workbook.save(OUTPUT)
    print(OUTPUT)


def create_quick_read_write_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("QuickReadWrite")
    sheet["A1"] = "openpyxlwings sample"
    sheet["A2"] = "This sheet is for read_range, read_cell_at, and write_values_at."
    sheet["A4"] = "Name"
    sheet["B4"] = "Score"
    sheet["C4"] = "Team"
    sheet["D4"] = "Updated By Python"
    rows = [
        ["Alice", 95, "Sales", None],
        ["Bob", 88, "Marketing", None],
        ["Charlie", 91, "Operations", None],
    ]
    for row_index, row in enumerate(rows, start=5):
        for column_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)

    style_title(sheet, "A1:D1")
    style_header(sheet, "A4:D4")
    apply_grid(sheet, 4, 1, 7, 4)
    set_widths(sheet, [18, 12, 18, 20])
    sheet.freeze_panes = "A5"


def create_bordered_table_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("BorderedTable")
    sheet["A1"] = "罫線テーブル編集サンプル"
    sheet["A2"] = "C5など、表内のセルを起点に get_bordered_table() を試せます。"

    values = [
        [None, "2026", "2027", "2028"],
        ["Region", "Sales", "Sales", "Plan"],
        ["East", 120, 135, 150],
        ["West", 98, 105, 118],
        ["North", 76, 82, 90],
    ]
    start_row = 4
    start_column = 2
    for row_offset, row in enumerate(values):
        for column_offset, value in enumerate(row):
            cell = sheet.cell(
                row=start_row + row_offset,
                column=start_column + column_offset,
                value=value,
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

    style_title(sheet, "A1:F1")
    sheet["B11"] = "Manual test:"
    sheet["B12"] = 'table = book.get_bordered_table("BorderedTable", row=5, column=3, header_rows=2, header_columns=1)'
    sheet["B13"] = "table.add_row([88, 99, 111], row_headers=['South']); table.save()"
    sheet["B12"].font = Font(name="Consolas", size=10)
    sheet["B13"].font = Font(name="Consolas", size=10)

    style_header(sheet, "B4:E5")
    apply_grid(sheet, 4, 2, 8, 5)
    set_widths(sheet, [4, 16, 14, 14, 14, 4])
    sheet.freeze_panes = "B6"

    chart = BarChart()
    chart.title = "Sales by Region"
    chart.y_axis.title = "Value"
    chart.x_axis.title = "Region"
    data = Reference(sheet, min_col=3, max_col=5, min_row=5, max_row=8)
    cats = Reference(sheet, min_col=2, min_row=6, max_row=8)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    sheet.add_chart(chart, "G4")


def create_broken_table_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("BrokenBorder")
    sheet["A1"] = "壊れた罫線テーブル例"
    sheet["A2"] = "内部罫線が欠けているため、get_bordered_table() は例外を出す想定です。"
    values = [["Name", "Q1"], ["Alice", 10], ["Bob", 20]]
    for row_offset, row in enumerate(values, start=4):
        for column_offset, value in enumerate(row, start=2):
            sheet.cell(row=row_offset, column=column_offset, value=value)

    thin = Side(style="thin", color="111827")
    full = Border(top=thin, bottom=thin, left=thin, right=thin)
    for row in range(4, 7):
        for column in range(2, 4):
            sheet.cell(row=row, column=column).border = full

    sheet.cell(row=5, column=2).border = Border(top=thin, bottom=thin, left=thin)
    sheet.cell(row=5, column=3).border = Border(top=thin, bottom=thin, right=thin)
    style_title(sheet, "A1:D1")
    style_header(sheet, "B4:C4")
    set_widths(sheet, [4, 18, 12, 30])


def style_title(sheet, address: str) -> None:
    for row in sheet[address]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True, size=13)
            cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 24


def style_header(sheet, address: str) -> None:
    for row in sheet[address]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_grid(sheet, start_row: int, start_column: int, end_row: int, end_column: int) -> None:
    thin = Side(style="thin", color="111827")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for row in range(start_row, end_row + 1):
        for column in range(start_column, end_column + 1):
            sheet.cell(row=row, column=column).border = border


def set_widths(sheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


if __name__ == "__main__":
    main()
