"""Border-based table detection and editing."""

from __future__ import annotations

from collections.abc import Callable, Iterator
from dataclasses import dataclass, field
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from openpyxlwings.exceptions import (
    BorderTableNotFoundError,
    BorderTableShapeError,
)

if TYPE_CHECKING:
    from openpyxlwings.workbook import CellValue, ExcelWorkbook, Table


@dataclass
class _Insertion:
    axis: str
    index: int


@dataclass
class BorderTable:
    """Editable model for a plain Excel range divided by borders.

    The table holds either the full rectangle of a bordered range
    (``partial=False``) or a subset of its columns selected by header
    (``partial=True``). Values are stored column-major: ``columns[i]`` is the
    i-th held column read from the table's first row down, header rows
    included. ``source_columns[i]`` is the Excel column the values came from,
    or ``None`` for a column added in memory that has no Excel origin yet.

    Saving a full table rewrites the whole rectangle (replaying row/column
    insertions at their recorded positions). Saving a partial table writes
    only the held columns back to their source columns; rows added in memory
    are inserted into the sheet, while added columns are appended to the right
    edge of the original table regardless of their virtual position.
    """

    workbook: ExcelWorkbook
    sheet: str | None
    start_row: int
    start_column: int
    columns: Table
    header_rows: int = 1
    header_columns: int = 0
    source_columns: list[int | None] | None = None
    partial: bool = False
    detected_end_row: int | None = None
    detected_end_column: int | None = None
    source_rows: list[int | None] | None = None
    partial_axis: str | None = None
    _original_rows: int = field(init=False, repr=False)
    _original_columns: int = field(init=False, repr=False)
    _insertions: list[_Insertion] = field(default_factory=list, init=False, repr=False)

    def __post_init__(self) -> None:
        if self.source_columns is None:
            self.source_columns = list(
                range(self.start_column, self.start_column + len(self.columns))
            )
        if self.source_rows is None:
            self.source_rows = list(range(self.start_row, self.start_row + self.row_count))
        if self.partial:
            if self.partial_axis is None:
                self.partial_axis = "column"
            if self.partial_axis not in ("column", "row"):
                raise BorderTableShapeError('partial_axis must be "column" or "row".')
            if self.detected_end_row is None or self.detected_end_column is None:
                raise BorderTableShapeError(
                    "partial tables require detected_end_row and detected_end_column."
                )
            self._pad_columns()
        self._validate_shape()
        self._original_rows = self.row_count
        self._original_columns = self.column_count

    @property
    def row_count(self) -> int:
        return len(self.columns[0]) if self.columns else 0

    @property
    def column_count(self) -> int:
        return len(self.columns)

    @property
    def added_rows(self) -> int:
        return max(0, self.row_count - self._original_rows)

    @property
    def added_columns(self) -> int:
        return sum(1 for source in self.source_columns if source is None)

    @property
    def end_row(self) -> int:
        if self.partial:
            return self.detected_end_row + self.added_rows
        return self.start_row + self.row_count - 1

    @property
    def end_column(self) -> int:
        if self.partial:
            return self.detected_end_column + self.added_columns
        return self.start_column + self.column_count - 1

    @property
    def range(self) -> str:
        return f"{_cell_address(self.start_row, self.start_column)}:{_cell_address(self.end_row, self.end_column)}"

    @property
    def data(self) -> Table:
        """The body area as a list of columns (column-major)."""

        return [column[self.header_rows :] for column in self.columns[self.header_columns :]]

    @property
    def row_headers(self) -> Table:
        if self.header_columns == 0:
            return []
        return [
            [column[row] for column in self.columns[: self.header_columns]]
            for row in range(self.header_rows, self.row_count)
        ]

    @property
    def column_headers(self) -> Table:
        return [
            [column[row] for column in self.columns[self.header_columns :]]
            for row in range(self.header_rows)
        ]

    def set_value(self, row: int, column: int, value: CellValue) -> None:
        """Set a value in the full table using 1-based table coordinates."""

        self._validate_table_position(row, column)
        self.columns[column - 1][row - 1] = value

    def set_body_value(self, row: int, column: int, value: CellValue) -> None:
        """Set a value in the body area using 1-based body coordinates."""

        body_rows = self.row_count - self.header_rows
        body_columns = self.column_count - self.header_columns
        _validate_position(row, column, max_row=body_rows, max_column=body_columns)
        self.columns[self.header_columns + column - 1][self.header_rows + row - 1] = value

    def find_body_row(self, row_header: CellValue | list[CellValue] | tuple[CellValue, ...]) -> int:
        """Find a body row by row header and return its 1-based body row index."""

        expected = self._normalize_row_header(row_header)
        matches = [
            index
            for index, actual in enumerate(self.row_headers, start=1)
            if tuple(actual) == expected
        ]
        if not matches:
            raise BorderTableShapeError("row_header was not found.")
        if len(matches) > 1:
            raise BorderTableShapeError("row_header matches multiple body rows.")
        return matches[0]

    def set_body_row_by_header(
        self,
        row_header: CellValue | list[CellValue] | tuple[CellValue, ...],
        values: list[CellValue],
    ) -> None:
        """Replace one body row selected by row header."""

        body_columns = self.column_count - self.header_columns
        if len(values) != body_columns:
            raise BorderTableShapeError("row values length does not match table body width.")

        row = self.find_body_row(row_header)
        table_row_index = self.header_rows + row - 1
        for column, value in zip(self.columns[self.header_columns :], values, strict=True):
            column[table_row_index] = value

    def select_columns_by_row(
        self,
        row_header: CellValue | list[CellValue] | tuple[CellValue, ...],
        condition: Callable[[CellValue], bool] | CellValue,
        *,
        match_case: bool = False,
    ) -> BorderTable:
        """Return a partial table of the body columns matching ``condition``.

        The row used for the comparison is selected by row header, like
        :meth:`find_body_row`. ``condition`` is either a callable that
        receives each column's raw value on that row, or a plain value that
        is compared like header matching (whitespace-stripped and, unless
        ``match_case`` is set, case-insensitive). Row-header columns are
        always kept.

        Values are copied, so editing the returned table does not affect
        this one. Saving the returned table writes only the selected columns
        back; derive it from a freshly detected (or saved) table so its
        bounds match the sheet.
        """

        row = self.find_body_row(row_header)
        table_row_index = self.header_rows + row - 1

        matches = _build_matcher(condition, match_case=match_case)

        kept_indexes = list(range(self.header_columns))
        kept_indexes.extend(
            index
            for index in range(self.header_columns, self.column_count)
            if matches(self.columns[index][table_row_index])
        )
        if len(kept_indexes) == self.header_columns:
            raise BorderTableShapeError("no body columns matched the condition.")

        return BorderTable(
            workbook=self.workbook,
            sheet=self.sheet,
            start_row=self.start_row,
            start_column=self.start_column,
            columns=[list(self.columns[index]) for index in kept_indexes],
            header_rows=self.header_rows,
            header_columns=self.header_columns,
            source_columns=[self.source_columns[index] for index in kept_indexes],
            partial=True,
            partial_axis="column",
            detected_end_row=self.end_row,
            detected_end_column=self.end_column,
        )

    def select_columns_by_column_header(
        self,
        condition: Callable[[CellValue], bool] | CellValue,
        *,
        match_case: bool = False,
    ) -> BorderTable:
        """Return a partial table of the body columns whose header matches.

        ``condition`` receives each body column's header value: a scalar when
        there is a single header row, otherwise a top-to-bottom tuple of the
        column's header values. A plain (non-callable) value is compared like
        header matching (whitespace-stripped and, unless ``match_case`` is set,
        case-insensitive). Row-header columns are always kept.

        The result is a column subset, editable and savable like
        :meth:`select_columns_by_row`.
        """

        if self.partial and self.partial_axis == "row":
            raise BorderTableShapeError("cannot column-filter a row-partial table.")

        matches = _build_matcher(condition, match_case=match_case)

        def header_of(index: int) -> CellValue | tuple[CellValue, ...]:
            if self.header_rows == 1:
                return self.columns[index][0]
            return tuple(self.columns[index][row] for row in range(self.header_rows))

        kept_indexes = list(range(self.header_columns))
        kept_indexes.extend(
            index
            for index in range(self.header_columns, self.column_count)
            if matches(header_of(index))
        )
        if len(kept_indexes) == self.header_columns:
            raise BorderTableShapeError("no body columns matched the condition.")

        return BorderTable(
            workbook=self.workbook,
            sheet=self.sheet,
            start_row=self.start_row,
            start_column=self.start_column,
            columns=[list(self.columns[index]) for index in kept_indexes],
            header_rows=self.header_rows,
            header_columns=self.header_columns,
            source_columns=[self.source_columns[index] for index in kept_indexes],
            source_rows=list(self.source_rows),
            partial=True,
            partial_axis="column",
            detected_end_row=self.end_row,
            detected_end_column=self.end_column,
        )

    def select_rows_by_row_header(
        self,
        condition: Callable[[CellValue], bool] | CellValue,
        *,
        match_case: bool = False,
    ) -> BorderTable:
        """Return a partial table of the body rows whose row header matches.

        ``condition`` receives each body row's row-header value: a scalar when
        there is a single row-header column, otherwise a left-to-right tuple of
        the row's header values. A plain (non-callable) value is compared like
        header matching (whitespace-stripped and, unless ``match_case`` is set,
        case-insensitive). Column-header rows are always kept.

        The result is a row subset: only the matched rows are held, and saving
        writes them back to their original Excel rows, leaving unmatched rows
        untouched. ``add_row`` / ``add_column`` and ``save`` are supported.
        """

        if self.header_columns == 0:
            raise BorderTableShapeError("table has no row headers.")
        if self.partial and self.partial_axis == "column":
            raise BorderTableShapeError("cannot row-filter a column-partial table.")

        matches = _build_matcher(condition, match_case=match_case)

        def header_of(row: int) -> CellValue | tuple[CellValue, ...]:
            if self.header_columns == 1:
                return self.columns[0][row]
            return tuple(self.columns[column][row] for column in range(self.header_columns))

        kept_rows = list(range(self.header_rows))
        kept_rows.extend(
            row
            for row in range(self.header_rows, self.row_count)
            if matches(header_of(row))
        )
        if len(kept_rows) == self.header_rows:
            raise BorderTableShapeError("no body rows matched the condition.")

        return BorderTable(
            workbook=self.workbook,
            sheet=self.sheet,
            start_row=self.start_row,
            start_column=self.start_column,
            columns=[[column[row] for row in kept_rows] for column in self.columns],
            header_rows=self.header_rows,
            header_columns=self.header_columns,
            source_columns=list(self.source_columns),
            source_rows=[self.source_rows[row] for row in kept_rows],
            partial=True,
            partial_axis="row",
            detected_end_row=self.end_row,
            detected_end_column=self.end_column,
        )

    def add_row(
        self,
        values: list[CellValue],
        *,
        row_headers: list[CellValue] | None = None,
        position: int | None = None,
    ) -> None:
        """Add a body row. ``position`` is 1-based inside the body."""

        body_rows = self.row_count - self.header_rows
        body_columns = self.column_count - self.header_columns
        insert_body_index = body_rows + 1 if position is None else position
        if insert_body_index < 1 or insert_body_index > body_rows + 1:
            raise BorderTableShapeError("row position is outside the body range.")
        if len(values) != body_columns:
            raise BorderTableShapeError("row values length does not match table body width.")

        headers = [] if row_headers is None else list(row_headers)
        if len(headers) != self.header_columns:
            raise BorderTableShapeError("row_headers length does not match header_columns.")

        appending = insert_body_index == body_rows + 1
        table_row_index = self.header_rows + insert_body_index - 1
        full_row = headers + list(values)
        for column, value in zip(self.columns, full_row, strict=True):
            column.insert(table_row_index, value)
        self.source_rows.insert(table_row_index, None)
        if self.partial and self.partial_axis == "row":
            # Row subset: the new row is the subset axis, placed at the table's
            # bottom edge on save via its ``None`` source row (no sheet insert).
            return
        if self.partial and appending:
            # The held grid can be shorter than the detected table, so appended
            # rows go below the original table, past rows inserted earlier.
            prior_rows = sum(1 for insertion in self._insertions if insertion.axis == "row")
            excel_row = self.detected_end_row + 1 + prior_rows
        else:
            excel_row = self.start_row + table_row_index
        self._insertions.append(_Insertion("row", excel_row))

    def add_column(
        self,
        values: list[CellValue],
        *,
        column_headers: list[CellValue] | None = None,
        position: int | None = None,
    ) -> None:
        """Add a body column. ``position`` is 1-based inside the body."""

        body_rows = self.row_count - self.header_rows
        body_columns = self.column_count - self.header_columns
        insert_body_index = body_columns + 1 if position is None else position
        if insert_body_index < 1 or insert_body_index > body_columns + 1:
            raise BorderTableShapeError("column position is outside the body range.")
        if len(values) != body_rows:
            raise BorderTableShapeError("column values length does not match table body height.")

        headers = [] if column_headers is None else list(column_headers)
        if len(headers) != self.header_rows:
            raise BorderTableShapeError("column_headers length does not match header_rows.")

        table_column_index = self.header_columns + insert_body_index - 1
        new_column = headers + list(values)
        self.columns.insert(table_column_index, new_column)
        self.source_columns.insert(table_column_index, None)
        if not self.partial:
            # Partial tables append new columns to the table's right edge on
            # save; the virtual position only affects in-memory ordering.
            self._insertions.append(
                _Insertion("column", self.start_column + table_column_index)
            )

    def add_header_row(
        self,
        values: list[CellValue],
        *,
        position: int | None = None,
    ) -> None:
        """Add a column-header row. ``position`` is 1-based inside headers."""

        if self.partial:
            raise BorderTableShapeError("header rows cannot be added to a partial table.")
        insert_index = self.header_rows + 1 if position is None else position
        if insert_index < 1 or insert_index > self.header_rows + 1:
            raise BorderTableShapeError("header row position is outside the header range.")
        if len(values) != self.column_count:
            raise BorderTableShapeError("header row values length does not match table width.")

        table_row_index = insert_index - 1
        for column, value in zip(self.columns, values, strict=True):
            column.insert(table_row_index, value)
        self.source_rows.insert(table_row_index, None)
        self.header_rows += 1
        self._insertions.append(_Insertion("row", self.start_row + table_row_index))

    def add_header_column(
        self,
        values: list[CellValue],
        *,
        position: int | None = None,
    ) -> None:
        """Add a row-header column. ``position`` is 1-based inside headers."""

        if self.partial:
            raise BorderTableShapeError("header columns cannot be added to a partial table.")
        insert_index = self.header_columns + 1 if position is None else position
        if insert_index < 1 or insert_index > self.header_columns + 1:
            raise BorderTableShapeError("header column position is outside the header range.")
        if len(values) != self.row_count:
            raise BorderTableShapeError("header column values length does not match table height.")

        table_column_index = insert_index - 1
        self.columns.insert(table_column_index, list(values))
        self.source_columns.insert(table_column_index, None)
        self.header_columns += 1
        self._insertions.append(_Insertion("column", self.start_column + table_column_index))

    def save(self, path: str | Path | None = None) -> None:
        """Write the edited table back to its workbook position.

        With ``path`` the edit is saved to a separate file, leaving the
        original file unchanged. As with a spreadsheet "Save As", the session's
        working workbook then becomes that new file, so later ``save()`` calls
        without a path also target it.
        """

        self.workbook._save_bordered_table(self, path)
        if self.partial and self.partial_axis == "row":
            # Rebaseline the row subset: appended rows now exist at the bottom
            # edge and appended columns at the right edge, so a second save does
            # not re-insert them.
            next_row = self.detected_end_row
            for index, source in enumerate(self.source_rows):
                if source is None:
                    next_row += 1
                    self.source_rows[index] = next_row
            self.detected_end_row = next_row
            next_column = self.detected_end_column
            for index, source in enumerate(self.source_columns):
                if source is None:
                    next_column += 1
                    self.source_columns[index] = next_column
            self.detected_end_column = next_column
        elif self.partial:
            # Column subset: appended columns now exist at the table's right edge.
            self.detected_end_row += self.added_rows
            next_column = self.detected_end_column
            for index, source in enumerate(self.source_columns):
                if source is None:
                    next_column += 1
                    self.source_columns[index] = next_column
            self.detected_end_column = next_column
        self._original_rows = self.row_count
        self._original_columns = self.column_count
        self._insertions.clear()

    def _pad_columns(self) -> None:
        height = max((len(column) for column in self.columns), default=0)
        for column in self.columns:
            if len(column) < height:
                column.extend([None] * (height - len(column)))

    def _validate_shape(self) -> None:
        if not self.columns:
            raise BorderTableShapeError("table values cannot be empty.")
        height = len(self.columns[0])
        if height == 0:
            raise BorderTableShapeError("table height cannot be zero.")
        if any(len(column) != height for column in self.columns):
            raise BorderTableShapeError("table values must be rectangular.")
        if len(self.source_columns) != len(self.columns):
            raise BorderTableShapeError("source_columns length does not match columns.")
        if len(self.source_rows) != height:
            raise BorderTableShapeError("source_rows length does not match rows.")
        if self.header_rows < 0 or self.header_columns < 0:
            raise BorderTableShapeError("header counts cannot be negative.")
        if self.header_rows >= height:
            raise BorderTableShapeError("header_rows must leave at least one body row.")
        if self.partial:
            if self.header_columns > len(self.columns):
                raise BorderTableShapeError(
                    "header_columns cannot exceed the number of held columns."
                )
        elif self.header_columns >= len(self.columns):
            raise BorderTableShapeError("header_columns must leave at least one body column.")

    def _validate_table_position(self, row: int, column: int) -> None:
        _validate_position(row, column, max_row=self.row_count, max_column=self.column_count)

    def _normalize_row_header(
        self,
        row_header: CellValue | list[CellValue] | tuple[CellValue, ...],
    ) -> tuple[CellValue, ...]:
        if self.header_columns == 0:
            raise BorderTableShapeError("table has no row headers.")

        headers = (
            tuple(row_header)
            if isinstance(row_header, (list, tuple))
            else (row_header,)
        )
        if len(headers) != self.header_columns:
            raise BorderTableShapeError("row_header length does not match header_columns.")
        return headers


def detect_table_region(
    workbook: ExcelWorkbook,
    worksheet: Worksheet,
    sheet: str | None,
    row: int,
    column: int,
    *,
    header_rows: int = 1,
    header_columns: int = 0,
) -> BorderTable:
    """Detect a table from its top-left cell by growing right and down.

    Unlike border-based detection this accepts tables that are values only,
    borders only, or bordered with gaps: a cell belongs to the region when it
    has a value or any border. The region grows column by column and row by
    row until it hits a row/column with no content inside the current span.
    A lone top border on the next row (or a lone left border on the next
    column) counts as the table's closing line, not as content.
    """

    _validate_position(row, column)

    _, _, end_row, end_column = _grow_region(worksheet, row, column, row, column)
    if end_row == row and end_column == column and not _cell_has_content(worksheet, row, column):
        raise BorderTableNotFoundError("A table was not found at the start cell.")

    columns = [
        [cell.value for cell in column_cells]
        for column_cells in worksheet.iter_cols(
            min_row=row,
            max_row=end_row,
            min_col=column,
            max_col=end_column,
        )
    ]
    return BorderTable(
        workbook=workbook,
        sheet=sheet,
        start_row=row,
        start_column=column,
        columns=columns,
        header_rows=header_rows,
        header_columns=header_columns,
    )


def _grow_region(
    worksheet: Worksheet,
    start_row: int,
    start_column: int,
    end_row: int,
    end_column: int,
    *,
    grow_left: bool = False,
) -> tuple[int, int, int, int]:
    """Grow the initial rectangle over adjacent content and return its bounds.

    Growth goes right and down (and left when ``grow_left`` is set) until no
    direction extends, so an L-shaped run of content still yields its full
    bounding box. The top edge never moves: callers know the table's first
    row (the anchor itself, or derived from ``header_rows``).
    """

    max_row = worksheet.max_row
    max_column = worksheet.max_column

    changed = True
    while changed:
        changed = False
        while (
            grow_left
            and start_column > 1
            and _column_extends_region_left(worksheet, start_column - 1, start_row, end_row)
        ):
            start_column -= 1
            changed = True
        while end_column < max_column and _column_extends_region(
            worksheet, end_column + 1, start_row, end_row
        ):
            end_column += 1
            changed = True
        while end_row < max_row and _row_extends_region(
            worksheet, end_row + 1, start_column, end_column
        ):
            end_row += 1
            changed = True

    return start_row, start_column, end_row, end_column


def _row_extends_region(
    worksheet: Worksheet,
    row: int,
    start_column: int,
    end_column: int,
) -> bool:
    """Whether the candidate row below the region carries table content.

    A cell whose only border is a top border is treated as the closing line
    of the row above, so it does not pull the candidate row into the table.
    """

    for column in range(start_column, end_column + 1):
        cell = worksheet.cell(row=row, column=column)
        if cell.value is not None:
            return True
        border = cell.border
        if _has_side(border.bottom) or _has_side(border.left) or _has_side(border.right):
            return True
    return False


def _column_extends_region(
    worksheet: Worksheet,
    column: int,
    start_row: int,
    end_row: int,
) -> bool:
    """Whether the candidate column right of the region carries table content.

    A cell whose only border is a left border is treated as the closing line
    of the column to the left, so it does not pull the candidate column into
    the table.
    """

    for row in range(start_row, end_row + 1):
        cell = worksheet.cell(row=row, column=column)
        if cell.value is not None:
            return True
        border = cell.border
        if _has_side(border.right) or _has_side(border.top) or _has_side(border.bottom):
            return True
    return False


def _column_extends_region_left(
    worksheet: Worksheet,
    column: int,
    start_row: int,
    end_row: int,
) -> bool:
    """Whether the candidate column left of the region carries table content.

    A cell whose only border is a right border is treated as the closing line
    of the column to the right, so it does not pull the candidate column into
    the table.
    """

    for row in range(start_row, end_row + 1):
        cell = worksheet.cell(row=row, column=column)
        if cell.value is not None:
            return True
        border = cell.border
        if _has_side(border.left) or _has_side(border.top) or _has_side(border.bottom):
            return True
    return False


def _cell_has_content(worksheet: Worksheet, row: int, column: int) -> bool:
    if worksheet.cell(row=row, column=column).value is not None:
        return True
    return _cell_has_any_border(worksheet, row, column)


def detect_bordered_table_by_header(
    workbook: ExcelWorkbook,
    worksheet: Worksheet,
    sheet: str | None,
    header_values: list[CellValue],
    *,
    value_header_contains: str,
    header_rows: int = 1,
    match_case: bool = False,
) -> BorderTable:
    """Detect a table by fixed header values and value-column text.

    Every cell run matching ``header_values`` is tried as a candidate. The
    table's first row is derived from ``header_rows`` (the match sits on the
    ``header_rows``-th table row), and the extent is found with the same
    content-based region growth as the top-left-cell search, so borders are
    not required. Candidates without a body row below the header are skipped.
    """

    if not header_values:
        raise BorderTableShapeError("header_values must contain at least one value.")
    if header_rows < 1:
        raise BorderTableShapeError("header_rows must be 1 or greater.")
    if not value_header_contains:
        raise BorderTableShapeError("value_header_contains cannot be empty.")

    header_width = len(header_values)
    for row, column in _iter_sequence_anchor_cells(worksheet, header_values, match_case=match_case):
        top_row = row - header_rows + 1
        if top_row < 1:
            continue

        start_row, start_column, end_row, end_column = _grow_region(
            worksheet, top_row, column, row, column, grow_left=True
        )
        if end_row <= row:
            continue  # no body rows below the header row

        relative_header_row = [
            worksheet.cell(row=row, column=header_column).value
            for header_column in range(start_column, end_column + 1)
        ]
        first_value_column = _find_first_value_header_column(
            relative_header_row,
            value_header_contains,
            match_case=match_case,
        )
        if first_value_column is None:
            continue
        if first_value_column <= header_width:
            raise BorderTableShapeError(
                "value-header columns must be to the right of header_values."
            )
        if first_value_column != header_width + 1:
            raise BorderTableShapeError(
                "header_values must cover every row-header column before the value area."
            )

        columns = [
            [cell.value for cell in column_cells]
            for column_cells in worksheet.iter_cols(
                min_row=start_row,
                max_row=end_row,
                min_col=start_column,
                max_col=end_column,
            )
        ]
        try:
            return BorderTable(
                workbook=workbook,
                sheet=sheet,
                start_row=start_row,
                start_column=start_column,
                columns=columns,
                header_rows=header_rows,
                header_columns=first_value_column - 1,
            )
        except BorderTableShapeError:
            continue  # malformed candidate; keep searching

    raise BorderTableNotFoundError("A table matching the header values was not found.")


def detect_bordered_table_by_columns(
    workbook: ExcelWorkbook,
    worksheet: Worksheet,
    sheet: str | None,
    header_values: list[CellValue],
    *,
    value_header_contains: str | None = None,
    header_rows: int = 1,
    match_case: bool = False,
) -> BorderTable:
    """Detect a table and hold only a subset of its columns.

    ``header_values`` are matched exactly (order preserved) anywhere on the
    table's header row. When ``value_header_contains`` is given, every header
    cell that contains the text is also selected (left to right). The table
    extent is found with the same content-based region growth as the
    top-left-cell search, so borders are not required. Each selected column
    is read from the table's first row down; below the header the read
    continues while a cell has a value, a top border, or a bottom border, and
    ragged columns are squared up with ``None``.
    """

    if not header_values:
        raise BorderTableShapeError("header_values must contain at least one value.")
    if header_rows < 1:
        raise BorderTableShapeError("header_rows must be 1 or greater.")

    for row, column in _iter_value_anchor_cells(worksheet, header_values[0], match_case=match_case):
        top_row = row - header_rows + 1
        if top_row < 1:
            continue

        start_row, start_column, end_row, end_column = _grow_region(
            worksheet, top_row, column, row, column, grow_left=True
        )
        if end_row <= row:
            continue  # no body rows below the header row

        selected = _select_source_columns(
            worksheet,
            row,
            start_column,
            end_column,
            header_values,
            value_header_contains,
            match_case=match_case,
        )
        if selected is None:
            continue

        columns = []
        for source in selected:
            values = [
                worksheet.cell(row=header_area_row, column=source).value
                for header_area_row in range(start_row, row + 1)
            ]
            values.extend(_read_column_body(worksheet, row + 1, source, max_row=end_row))
            columns.append(values)

        try:
            return BorderTable(
                workbook=workbook,
                sheet=sheet,
                start_row=start_row,
                start_column=start_column,
                columns=columns,
                header_rows=header_rows,
                header_columns=len(header_values),
                source_columns=list(selected),
                partial=True,
                detected_end_row=end_row,
                detected_end_column=end_column,
            )
        except BorderTableShapeError:
            continue  # malformed candidate; keep searching

    raise BorderTableNotFoundError(
        "A table matching the requested column headers was not found."
    )


def _select_source_columns(
    worksheet: Worksheet,
    header_row: int,
    start_column: int,
    end_column: int,
    header_values: list[CellValue],
    value_header_contains: str | None,
    *,
    match_case: bool,
) -> list[int] | None:
    """Resolve the Excel columns selected by exact and substring header match."""

    used: set[int] = set()
    selected: list[int] = []

    for header_value in header_values:
        column = _find_header_column(
            worksheet,
            header_row,
            start_column,
            end_column,
            header_value,
            match_case=match_case,
            skip=used,
        )
        if column is None:
            return None
        used.add(column)
        selected.append(column)

    if value_header_contains:
        value_columns = [
            column
            for column in range(start_column, end_column + 1)
            if column not in used
            and _header_contains(
                worksheet.cell(row=header_row, column=column).value,
                value_header_contains,
                match_case=match_case,
            )
        ]
        if not value_columns:
            return None
        selected.extend(value_columns)

    return selected


def _read_column_body(
    worksheet: Worksheet,
    start_row: int,
    column: int,
    *,
    max_row: int | None = None,
) -> list[CellValue]:
    """Read a column downward while value/top-border/bottom-border continues."""

    if max_row is None:
        max_row = worksheet.max_row
    values: list[CellValue] = []
    row = start_row
    while row <= max_row:
        value = worksheet.cell(row=row, column=column).value
        if (
            value is not None
            or _has_top_border(worksheet, row, column)
            or _has_bottom_border(worksheet, row, column)
        ):
            values.append(value)
            row += 1
        else:
            break
    return values


def _find_header_column(
    worksheet: Worksheet,
    header_row: int,
    start_column: int,
    end_column: int,
    expected: CellValue,
    *,
    match_case: bool,
    skip: set[int] | None = None,
) -> int | None:
    target = _normalize_value(expected, match_case=match_case)
    for column in range(start_column, end_column + 1):
        if skip is not None and column in skip:
            continue
        actual = worksheet.cell(row=header_row, column=column).value
        if _normalize_value(actual, match_case=match_case) == target:
            return column
    return None


def _header_contains(value: CellValue, text: str, *, match_case: bool) -> bool:
    haystack = "" if value is None else str(value)
    needle = text
    if not match_case:
        haystack = haystack.casefold()
        needle = needle.casefold()
    return needle in haystack


def _iter_sequence_anchor_cells(
    worksheet: Worksheet,
    header_values: list[CellValue],
    *,
    match_case: bool,
) -> Iterator[tuple[int, int]]:
    """Yield every cell where ``header_values`` appear consecutively to the right."""

    width = len(header_values)
    for row in range(1, worksheet.max_row + 1):
        for column in range(1, worksheet.max_column - width + 2):
            if _header_values_match(worksheet, row, column, header_values, match_case=match_case):
                yield row, column


def _iter_value_anchor_cells(
    worksheet: Worksheet,
    expected: CellValue,
    *,
    match_case: bool,
) -> Iterator[tuple[int, int]]:
    """Yield every cell whose value matches ``expected`` exactly."""

    target = _normalize_value(expected, match_case=match_case)
    for row in range(1, worksheet.max_row + 1):
        for column in range(1, worksheet.max_column + 1):
            actual = worksheet.cell(row=row, column=column).value
            if _normalize_value(actual, match_case=match_case) == target:
                yield row, column


def _has_top_border(worksheet: Worksheet, row: int, column: int) -> bool:
    return _has_side(worksheet.cell(row=row, column=column).border.top)


def _has_bottom_border(worksheet: Worksheet, row: int, column: int) -> bool:
    return _has_side(worksheet.cell(row=row, column=column).border.bottom)


def _has_side(side: object) -> bool:
    return bool(getattr(side, "style", None))


def _cell_has_any_border(worksheet: Worksheet, row: int, column: int) -> bool:
    border = worksheet.cell(row=row, column=column).border
    return any(
        _has_side(side)
        for side in (border.top, border.bottom, border.left, border.right)
    )


def _header_values_match(
    worksheet: Worksheet,
    row: int,
    column: int,
    expected_values: list[CellValue],
    *,
    match_case: bool,
) -> bool:
    for offset, expected in enumerate(expected_values):
        actual = worksheet.cell(row=row, column=column + offset).value
        if _normalize_value(actual, match_case=match_case) != _normalize_value(
            expected,
            match_case=match_case,
        ):
            return False
    return True


def _find_first_value_header_column(
    values: list[CellValue],
    text: str,
    *,
    match_case: bool,
) -> int | None:
    needle = text if match_case else text.casefold()
    for index, value in enumerate(values, start=1):
        haystack = "" if value is None else str(value)
        if not match_case:
            haystack = haystack.casefold()
        if needle in haystack:
            return index
    return None


def _normalize_value(value: CellValue, *, match_case: bool) -> str:
    normalized = "" if value is None else str(value).strip()
    return normalized if match_case else normalized.casefold()


def _build_matcher(
    condition: Callable[[object], bool] | object,
    *,
    match_case: bool,
) -> Callable[[object], bool]:
    """Turn ``condition`` into a predicate over header/cell values.

    A callable is used as-is. A plain value becomes an equality test using the
    same normalization as header matching (whitespace-stripped and, unless
    ``match_case`` is set, case-insensitive). Tuple values (multi-header) are
    compared element-wise.
    """

    if callable(condition):
        return condition

    def normalize(value: object) -> object:
        if isinstance(value, tuple):
            return tuple(_normalize_value(item, match_case=match_case) for item in value)
        return _normalize_value(value, match_case=match_case)

    expected = normalize(condition)

    def matches(value: object) -> bool:
        return normalize(value) == expected

    return matches


def _cell_address(row: int, column: int) -> str:
    return f"{get_column_letter(column)}{row}"


def _validate_position(
    row: int,
    column: int,
    *,
    max_row: int | None = None,
    max_column: int | None = None,
) -> None:
    if row < 1:
        raise BorderTableShapeError("row must be 1 or greater.")
    if column < 1:
        raise BorderTableShapeError("column must be 1 or greater.")
    if max_row is not None and row > max_row:
        raise BorderTableShapeError("row is outside the table range.")
    if max_column is not None and column > max_column:
        raise BorderTableShapeError("column is outside the table range.")
