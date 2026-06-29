"""Unified Excel workbook API."""

from __future__ import annotations

from collections.abc import Sequence
from pathlib import Path
from typing import Any

import xlwings as xw
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from openpyxlwings.border_table import (
    BorderTable,
    detect_bordered_table,
    detect_bordered_table_by_header,
)
from openpyxlwings.exceptions import ExcelWriteError, SheetNotFoundError
from openpyxlwings.format import ExtractedMatch, TablePattern, extract_pattern

CellValue = str | int | float | bool | None
Table = list[list[CellValue]]
Scalar = str | int | float | bool | None
RangeValue = Scalar | Sequence[Scalar] | Sequence[Sequence[Scalar]]


class ExcelWorkbook:
    """Read with openpyxl and write with an isolated xlwings Excel instance.

    Reads are lazy and use openpyxl. Writes are also lazy and use a new Excel
    application instance by default, so this class does not operate on the
    user's already-open Excel windows.

    Reads default to a full (non read-only) load so that cell styles such as
    borders are available and the same cached workbook serves both value and
    bordered-table reads. Pass ``read_only=True`` to opt into openpyxl's
    memory-light streaming mode for value-only reads of very large files.
    """

    def __init__(
        self,
        path: str | Path,
        *,
        data_only: bool = True,
        read_only: bool = False,
        keep_vba: bool = False,
        visible: bool = False,
        update_links: bool = False,
        app: xw.App | None = None,
    ) -> None:
        self.path = Path(path)
        self.data_only = data_only
        self.read_only = read_only
        self.keep_vba = keep_vba
        self.visible = visible
        self.update_links = update_links
        self._reader = _OpenpyxlReadSession(
            self.path,
            data_only=data_only,
            read_only=read_only,
            keep_vba=keep_vba,
        )
        self._writer = _XlwingsWriteSession(
            self.path,
            visible=visible,
            update_links=update_links,
            app=app,
        )

    def __enter__(self) -> ExcelWorkbook:
        return self

    def __exit__(self, exc_type: object, _exc: object, _tb: object) -> None:
        self.close(save=exc_type is None)

    def close(self, *, save: bool = True) -> None:
        """Close any open read/write sessions."""

        self._reader.close()
        self._writer.close(save=save)

    def save(self, path: str | Path | None = None) -> None:
        """Save pending Excel-side writes."""

        self._writer.save(path)

    def sheet_names(self) -> list[str]:
        """Return workbook sheet names."""

        return self._reader.sheet_names()

    def read_cell(self, sheet: str | None, cell: str) -> CellValue:
        """Read one cell value."""

        return self._reader.read_cell(sheet, cell)

    def read_cell_at(self, sheet: str | None, row: int, column: int) -> CellValue:
        """Read one cell value by 1-based row and column numbers."""

        return self._reader.read_cell_at(sheet, row, column)

    def read_range_at(
        self,
        sheet: str | None,
        start_row: int,
        start_column: int,
        end_row: int,
        end_column: int,
    ) -> Table:
        """Read a rectangular range by 1-based row and column numbers."""

        return self._reader.read_range_at(
            sheet,
            start_row,
            start_column,
            end_row,
            end_column,
        )

    def read_range(self, sheet: str | None, address: str) -> Table:
        """Read a rectangular range such as ``A1:D10``."""

        return self._reader.read_range(sheet, address)

    def read_sheet(
        self,
        sheet: str | None = None,
        *,
        trim_empty: bool = True,
    ) -> Table:
        """Read the used area of a sheet as a list of rows."""

        return self._reader.read_sheet(sheet, trim_empty=trim_empty)

    def iter_rows(
        self,
        sheet: str | None = None,
        *,
        min_row: int | None = None,
        max_row: int | None = None,
        min_col: int | None = None,
        max_col: int | None = None,
    ):
        """Yield rows from a worksheet as tuples of values."""

        yield from self._reader.iter_rows(
            sheet,
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
        )

    def write_values(
        self,
        sheet: str | None,
        cell: str,
        values: RangeValue,
        *,
        expand: bool = False,
    ) -> None:
        """Write a scalar, row, column, or two-dimensional values at ``cell``."""

        self._reader.close()
        self._writer.write_values(sheet, cell, values, expand=expand)

    def write_values_at(
        self,
        sheet: str | None,
        row: int,
        column: int,
        values: RangeValue,
        *,
        expand: bool = False,
    ) -> None:
        """Write values by 1-based row and column numbers."""

        self._reader.close()
        self._writer.write_values_at(sheet, row, column, values, expand=expand)

    def clear_contents(
        self,
        sheet: str | None,
        address: str,
    ) -> None:
        """Clear values/formulas from a range without deleting shapes or formats."""

        self._reader.close()
        self._writer.clear_contents(sheet, address)

    def clear_contents_at(
        self,
        sheet: str | None,
        start_row: int,
        start_column: int,
        end_row: int,
        end_column: int,
    ) -> None:
        """Clear a range by 1-based row and column numbers."""

        self._reader.close()
        self._writer.clear_contents_at(
            sheet,
            start_row,
            start_column,
            end_row,
            end_column,
        )

    def get_bordered_table(
        self,
        sheet: str | None,
        row: int,
        column: int,
        *,
        header_rows: int = 1,
        header_columns: int = 0,
    ) -> BorderTable:
        """Detect a bordered table that contains ``row``/``column``."""

        worksheet = self._reader.styled_sheet(sheet)
        return detect_bordered_table(
            self,
            worksheet,
            sheet,
            row,
            column,
            header_rows=header_rows,
            header_columns=header_columns,
        )

    def get_bordered_table_by_header(
        self,
        sheet: str | None,
        header_values: list[CellValue],
        *,
        value_header_contains: str,
        header_row: int = 1,
        match_case: bool = False,
    ) -> BorderTable:
        """Detect a bordered table by header values and value-column text."""

        worksheet = self._reader.styled_sheet(sheet)
        return detect_bordered_table_by_header(
            self,
            worksheet,
            sheet,
            header_values,
            value_header_contains=value_header_contains,
            header_row=header_row,
            match_case=match_case,
        )

    def extract(
        self,
        pattern: TablePattern,
        *,
        sheets: list[str] | None = None,
        ranges: dict[str, str] | None = None,
    ) -> list[ExtractedMatch]:
        """Extract every table matching an Excel format pattern."""

        return extract_pattern(self, pattern, sheets=sheets, ranges=ranges)

    def _save_bordered_table(self, table: BorderTable) -> None:
        self._reader.close()
        self._writer.save_bordered_table(table)


class _OpenpyxlReadSession:
    def __init__(
        self,
        path: Path,
        *,
        data_only: bool,
        read_only: bool,
        keep_vba: bool,
    ) -> None:
        self.path = path
        self.data_only = data_only
        self.read_only = read_only
        self.keep_vba = keep_vba
        self._workbook = None
        self._styled_workbooks: dict[bool, Any] = {}

    @property
    def workbook(self):
        if self._workbook is None:
            self._workbook = load_workbook(
                self.path,
                read_only=self.read_only,
                data_only=self.data_only,
                keep_vba=self.keep_vba,
            )
        return self._workbook

    def styled_workbook_for(self, *, data_only: bool):
        # Border/style inspection requires a non-read-only workbook, since
        # openpyxl's read-only mode does not load cell styles. Cache one per
        # ``data_only`` flavor so repeated reads do not reopen the file.
        if not self.read_only and data_only == self.data_only:
            return self.workbook
        if data_only not in self._styled_workbooks:
            self._styled_workbooks[data_only] = load_workbook(
                self.path,
                read_only=False,
                data_only=data_only,
                keep_vba=self.keep_vba,
            )
        return self._styled_workbooks[data_only]

    @property
    def styled_workbook(self):
        return self.styled_workbook_for(data_only=self.data_only)

    def close(self) -> None:
        if self._workbook is not None:
            self._workbook.close()
            self._workbook = None
        for workbook in self._styled_workbooks.values():
            workbook.close()
        self._styled_workbooks.clear()

    def sheet_names(self) -> list[str]:
        return list(self.workbook.sheetnames)

    def sheet(self, name: str | None = None) -> Worksheet:
        return self._resolve_sheet(self.workbook, name)

    def styled_sheet(self, name: str | None = None) -> Worksheet:
        return self._resolve_sheet(self.styled_workbook, name)

    @staticmethod
    def _resolve_sheet(workbook, name: str | None) -> Worksheet:
        if name is None:
            return workbook.active
        if name not in workbook.sheetnames:
            raise SheetNotFoundError(f"Sheet not found: {name}")
        return workbook[name]

    def read_cell(self, sheet: str | None, cell: str) -> CellValue:
        return self.sheet(sheet)[cell].value

    def read_cell_at(self, sheet: str | None, row: int, column: int) -> CellValue:
        _validate_position(row, column)
        return self.sheet(sheet).cell(row=row, column=column).value

    def read_range(self, sheet: str | None, address: str) -> Table:
        worksheet = self.sheet(sheet)
        min_col, min_row, max_col, max_row = range_boundaries(address)
        return self._read_range_by_bounds(worksheet, min_row, min_col, max_row, max_col)

    def read_range_at(
        self,
        sheet: str | None,
        start_row: int,
        start_column: int,
        end_row: int,
        end_column: int,
    ) -> Table:
        _validate_range_position(start_row, start_column, end_row, end_column)
        worksheet = self.sheet(sheet)
        return self._read_range_by_bounds(
            worksheet,
            start_row,
            start_column,
            end_row,
            end_column,
        )

    def _read_range_by_bounds(
        self,
        worksheet: Worksheet,
        min_row: int,
        min_col: int,
        max_row: int,
        max_col: int,
    ) -> Table:
        rows = worksheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        )
        return [list(row) for row in rows]

    def read_sheet(
        self,
        sheet: str | None = None,
        *,
        trim_empty: bool = True,
    ) -> Table:
        rows = [list(row) for row in self.sheet(sheet).iter_rows(values_only=True)]
        if not trim_empty:
            return rows
        return _trim_empty_edges(rows)

    def iter_rows(
        self,
        sheet: str | None = None,
        *,
        min_row: int | None = None,
        max_row: int | None = None,
        min_col: int | None = None,
        max_col: int | None = None,
    ):
        yield from self.sheet(sheet).iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        )


class _XlwingsWriteSession:
    def __init__(
        self,
        path: Path,
        *,
        visible: bool,
        update_links: bool,
        app: xw.App | None,
    ) -> None:
        self.path = path
        self.visible = visible
        self.update_links = update_links
        self._external_app = app
        self._app: xw.App | None = None
        self._book: xw.Book | None = None

    @property
    def book(self) -> xw.Book:
        if self._book is None:
            self.open()
        if self._book is None:
            raise ExcelWriteError("Workbook is not open.")
        return self._book

    def open(self) -> None:
        if self._book is not None:
            return

        self._app = self._external_app or xw.App(
            visible=self.visible,
            add_book=False,
        )
        self._app.display_alerts = False
        self._app.screen_updating = False

        try:
            self._book = self._app.books.open(
                str(self.path.resolve()),
                update_links=self.update_links,
                read_only=False,
                notify=False,
                ignore_read_only_recommended=True,
            )
        except Exception as exc:  # pragma: no cover - depends on local Excel.
            self._cleanup_failed_open()
            raise ExcelWriteError(
                f"Could not open workbook for writing: {self.path}. "
                "If this workbook is already open in Excel, close it first or "
                "write to a copy to avoid touching the user's open session."
            ) from exc

        if _is_read_only_book(self._book):
            self.close(save=False)
            raise ExcelWriteError(
                f"Workbook opened read-only and was not modified: {self.path}. "
                "The file may already be open or locked by another Excel process."
            )

    def close(self, *, save: bool = True) -> None:
        book = self._book
        app = self._app
        self._book = None
        self._app = None

        try:
            if book is not None:
                if save:
                    book.save()
                book.close()
        finally:
            if self._external_app is None and app is not None:
                app.quit()

    def save(self, path: str | Path | None = None) -> None:
        if path is None:
            self.book.save()
        else:
            self.book.save(str(Path(path).resolve()))

    def sheet(self, name: str | None = None) -> xw.Sheet:
        if name is None:
            return self.book.sheets.active
        try:
            return self.book.sheets[name]
        except Exception as exc:  # pragma: no cover - xlwings API boundary.
            raise SheetNotFoundError(f"Sheet not found: {name}") from exc

    def write_values(
        self,
        sheet: str | None,
        cell: str,
        values: RangeValue,
        *,
        expand: bool = False,
    ) -> None:
        target = self.sheet(sheet).range(cell)
        if expand:
            target = target.expand()
        target.value = values

    def write_values_at(
        self,
        sheet: str | None,
        row: int,
        column: int,
        values: RangeValue,
        *,
        expand: bool = False,
    ) -> None:
        _validate_position(row, column)
        self.write_values(
            sheet,
            _cell_address(row, column),
            values,
            expand=expand,
        )

    def clear_contents(self, sheet: str | None, address: str) -> None:
        self.sheet(sheet).range(address).clear_contents()

    def clear_contents_at(
        self,
        sheet: str | None,
        start_row: int,
        start_column: int,
        end_row: int,
        end_column: int,
    ) -> None:
        _validate_range_position(start_row, start_column, end_row, end_column)
        self.clear_contents(
            sheet,
            _range_address(start_row, start_column, end_row, end_column),
        )

    def save_bordered_table(self, table: BorderTable) -> None:
        for insertion in table._insertions:
            if insertion.axis == "row":
                self.insert_row(table.sheet, insertion.index)
            else:
                self.insert_column(table.sheet, insertion.index)

        self.write_values_at(
            table.sheet,
            table.start_row,
            table.start_column,
            table.values,
        )
        self.apply_table_borders(
            table.sheet,
            table.start_row,
            table.start_column,
            table.end_row,
            table.end_column,
        )
        self.save()

    def insert_row(self, sheet: str | None, row: int) -> None:
        _validate_position(row, 1)
        self.sheet(sheet).api.Rows(row).Insert()

    def insert_column(self, sheet: str | None, column: int) -> None:
        _validate_position(1, column)
        self.sheet(sheet).api.Columns(column).Insert()

    def apply_table_borders(
        self,
        sheet: str | None,
        start_row: int,
        start_column: int,
        end_row: int,
        end_column: int,
    ) -> None:
        _validate_range_position(start_row, start_column, end_row, end_column)
        target = self.sheet(sheet).range(
            _range_address(start_row, start_column, end_row, end_column)
        )
        # Excel border constants: continuous line, thin weight.
        for border_index in (7, 8, 9, 10, 11, 12):
            border = target.api.Borders(border_index)
            border.LineStyle = 1
            border.Weight = 2

    def _cleanup_failed_open(self) -> None:
        app = self._app
        self._book = None
        self._app = None
        if self._external_app is None and app is not None:
            app.quit()


def sheet_names(path: str | Path) -> list[str]:
    """Return the sheet names for a workbook."""

    with ExcelWorkbook(path) as workbook:
        return workbook.sheet_names()


def read_range(path: str | Path, sheet: str | None, address: str) -> Table:
    """Read a rectangular range from a workbook."""

    with ExcelWorkbook(path) as workbook:
        return workbook.read_range(sheet, address)


def read_cell_at(path: str | Path, sheet: str | None, row: int, column: int) -> CellValue:
    """Read one cell by 1-based row and column numbers."""

    with ExcelWorkbook(path) as workbook:
        return workbook.read_cell_at(sheet, row, column)


def read_range_at(
    path: str | Path,
    sheet: str | None,
    start_row: int,
    start_column: int,
    end_row: int,
    end_column: int,
) -> Table:
    """Read a range by 1-based row and column numbers."""

    with ExcelWorkbook(path) as workbook:
        return workbook.read_range_at(
            sheet,
            start_row,
            start_column,
            end_row,
            end_column,
        )


def read_sheet(path: str | Path, sheet: str | None = None) -> Table:
    """Read all populated rows from a sheet."""

    with ExcelWorkbook(path) as workbook:
        return workbook.read_sheet(sheet)


def write_values(
    path: str | Path,
    sheet: str | None,
    cell: str,
    values: RangeValue,
    *,
    visible: bool = False,
) -> None:
    """Open a workbook in an isolated Excel instance, write values, and save."""

    with ExcelWorkbook(path, visible=visible) as workbook:
        workbook.write_values(sheet, cell, values)


def write_values_at(
    path: str | Path,
    sheet: str | None,
    row: int,
    column: int,
    values: RangeValue,
    *,
    visible: bool = False,
) -> None:
    """Open a workbook, write values by row/column numbers, and save."""

    with ExcelWorkbook(path, visible=visible) as workbook:
        workbook.write_values_at(sheet, row, column, values)


def write_range(
    path: str | Path,
    sheet: str | None,
    cell: str,
    values: RangeValue,
    *,
    visible: bool = False,
) -> None:
    """Alias for :func:`write_values`."""

    write_values(path, sheet, cell, values, visible=visible)


def write_range_at(
    path: str | Path,
    sheet: str | None,
    row: int,
    column: int,
    values: RangeValue,
    *,
    visible: bool = False,
) -> None:
    """Alias for :func:`write_values_at`."""

    write_values_at(path, sheet, row, column, values, visible=visible)


def _trim_empty_edges(rows: Sequence[Sequence[Any]]) -> Table:
    table = [list(row) for row in rows]
    while table and _is_empty_row(table[-1]):
        table.pop()

    if not table:
        return []

    last_col = 0
    for row in table:
        for index, value in enumerate(row, start=1):
            if value is not None:
                last_col = max(last_col, index)

    return [row[:last_col] for row in table]


def _is_empty_row(row: Sequence[Any]) -> bool:
    return all(value is None for value in row)


def _is_read_only_book(book: xw.Book) -> bool:
    try:
        return bool(book.api.ReadOnly)
    except Exception:
        return False


def _cell_address(row: int, column: int) -> str:
    _validate_position(row, column)
    return f"{get_column_letter(column)}{row}"


def _range_address(
    start_row: int,
    start_column: int,
    end_row: int,
    end_column: int,
) -> str:
    _validate_range_position(start_row, start_column, end_row, end_column)
    return f"{_cell_address(start_row, start_column)}:{_cell_address(end_row, end_column)}"


def _validate_position(row: int, column: int) -> None:
    if row < 1:
        raise ValueError("row must be 1 or greater.")
    if column < 1:
        raise ValueError("column must be 1 or greater.")


def _validate_range_position(
    start_row: int,
    start_column: int,
    end_row: int,
    end_column: int,
) -> None:
    _validate_position(start_row, start_column)
    _validate_position(end_row, end_column)
    if end_row < start_row:
        raise ValueError("end_row must be greater than or equal to start_row.")
    if end_column < start_column:
        raise ValueError("end_column must be greater than or equal to start_column.")
