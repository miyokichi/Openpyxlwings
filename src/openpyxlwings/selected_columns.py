"""Column-selected virtual table built from a bordered Excel table.

Unlike :class:`~openpyxlwings.border_table.BorderTable`, which reads and writes
the whole rectangle of a bordered table, this model lets the caller pick only a
subset of the table's columns by header. The selected columns are held as a
virtual table that supports row/column additions, and writing it back only
touches the selected columns (and any appended rows/columns); unselected
columns are left untouched.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import TYPE_CHECKING

from openpyxl.worksheet.worksheet import Worksheet

from openpyxlwings.border_table import (
    _has_bottom_border,
    _has_top_border,
    _normalize_value,
    detect_bordered_table,
)
from openpyxlwings.exceptions import (
    BorderTableNotFoundError,
    BorderTableShapeError,
)

if TYPE_CHECKING:
    from openpyxlwings.workbook import CellValue, ExcelWorkbook


@dataclass
class _SelectedColumn:
    """One column of the virtual table.

    ``source_column`` is the 1-based Excel column the data came from, or
    ``None`` for a column added in memory that has no Excel origin yet.
    """

    header: CellValue
    source_column: int | None
    values: list[CellValue]


@dataclass
class SelectedColumnsTable:
    """Editable virtual table holding a subset of a bordered table's columns."""

    workbook: ExcelWorkbook
    sheet: str | None
    start_row: int
    start_column: int
    end_row: int
    end_column: int
    header_row: int
    columns: list[_SelectedColumn]
    _original_rows: int = field(init=False, repr=False)

    def __post_init__(self) -> None:
        self._pad_columns()
        self._original_rows = self.row_count

    @property
    def body_start_row(self) -> int:
        return self.header_row + 1

    @property
    def row_count(self) -> int:
        return max((len(column.values) for column in self.columns), default=0)

    @property
    def column_count(self) -> int:
        return len(self.columns)

    @property
    def column_headers(self) -> list[CellValue]:
        return [column.header for column in self.columns]

    @property
    def data(self) -> list[list[CellValue]]:
        """The selected columns as a rectangular, None-padded grid of rows."""

        return [
            [column.values[row] for column in self.columns]
            for row in range(self.row_count)
        ]

    def set_value(self, row: int, column: int, value: CellValue) -> None:
        """Set a body value using 1-based virtual coordinates."""

        if row < 1 or row > self.row_count:
            raise BorderTableShapeError("row is outside the virtual table range.")
        if column < 1 or column > self.column_count:
            raise BorderTableShapeError("column is outside the virtual table range.")
        self.columns[column - 1].values[row - 1] = value

    def add_row(
        self,
        values: list[CellValue],
        *,
        position: int | None = None,
    ) -> None:
        """Add a body row. ``position`` is 1-based inside the body."""

        if len(values) != self.column_count:
            raise BorderTableShapeError("row values length does not match column count.")
        insert_index = self.row_count if position is None else position - 1
        if insert_index < 0 or insert_index > self.row_count:
            raise BorderTableShapeError("row position is outside the body range.")
        for column, value in zip(self.columns, values, strict=True):
            column.values.insert(insert_index, value)

    def add_column(
        self,
        values: list[CellValue],
        *,
        header: CellValue = None,
        position: int | None = None,
    ) -> None:
        """Add a virtual column. ``position`` is 1-based among columns.

        The new column has no Excel origin, so it is written into a freshly
        inserted column at the end of the table when the table is saved.
        """

        if len(values) != self.row_count:
            raise BorderTableShapeError("column values length does not match row count.")
        insert_index = self.column_count if position is None else position - 1
        if insert_index < 0 or insert_index > self.column_count:
            raise BorderTableShapeError("column position is outside the table range.")
        self.columns.insert(
            insert_index,
            _SelectedColumn(header=header, source_column=None, values=list(values)),
        )

    @property
    def added_rows(self) -> int:
        return max(0, self.row_count - self._original_rows)

    def save(self) -> None:
        """Write the edited virtual table back to its workbook."""

        self.workbook._save_selected_columns_table(self)
        self._original_rows = self.row_count

    def _pad_columns(self) -> None:
        height = self.row_count
        for column in self.columns:
            if len(column.values) < height:
                column.values.extend([None] * (height - len(column.values)))


def detect_selected_columns_table(
    workbook: ExcelWorkbook,
    worksheet: Worksheet,
    sheet: str | None,
    header_values: list[CellValue],
    *,
    value_header_contains: str | None = None,
    header_row: int = 1,
    match_case: bool = False,
    require_inner_borders: bool = True,
) -> SelectedColumnsTable:
    """Detect a bordered table and select a subset of its columns by header.

    ``header_values`` are matched exactly (order preserved). When
    ``value_header_contains`` is given, every header cell that contains the text
    is also selected (left to right). Only the selected columns' data is read,
    each one walked downward while a cell has a value, a top border, or a bottom
    border.
    """

    if not header_values:
        raise BorderTableShapeError("header_values must contain at least one value.")
    if header_row < 1:
        raise BorderTableShapeError("header_row must be 1 or greater.")

    for row in range(1, worksheet.max_row + 1):
        anchor_column = _find_header_column(
            worksheet,
            row,
            1,
            worksheet.max_column,
            header_values[0],
            match_case=match_case,
        )
        if anchor_column is None:
            continue

        try:
            table = detect_bordered_table(
                workbook,
                worksheet,
                sheet,
                row,
                anchor_column,
                header_rows=header_row,
                header_columns=0,
                require_inner_borders=require_inner_borders,
            )
        except (BorderTableNotFoundError, BorderTableShapeError):
            continue

        if row - table.start_row + 1 != header_row:
            continue

        columns = _select_columns(
            worksheet,
            row,
            table.start_column,
            table.end_column,
            header_values,
            value_header_contains,
            match_case=match_case,
        )
        if columns is None:
            continue

        return SelectedColumnsTable(
            workbook=workbook,
            sheet=sheet,
            start_row=table.start_row,
            start_column=table.start_column,
            end_row=table.end_row,
            end_column=table.end_column,
            header_row=row,
            columns=columns,
        )

    raise BorderTableNotFoundError(
        "A bordered table matching the requested column headers was not found."
    )


def _select_columns(
    worksheet: Worksheet,
    header_row: int,
    start_column: int,
    end_column: int,
    header_values: list[CellValue],
    value_header_contains: str | None,
    *,
    match_case: bool,
) -> list[_SelectedColumn] | None:
    used: set[int] = set()
    selected: list[_SelectedColumn] = []

    for header_value in header_values:
        column = _find_header_column(
            worksheet,
            header_row,
            start_column,
            end_column,
            header_value,
            match_case=match_case,
            skip=used,
        )
        if column is None:
            return None
        used.add(column)
        selected.append(_build_column(worksheet, header_row, column))

    if value_header_contains:
        value_columns = [
            column
            for column in range(start_column, end_column + 1)
            if column not in used
            and _header_contains(
                worksheet.cell(row=header_row, column=column).value,
                value_header_contains,
                match_case=match_case,
            )
        ]
        if not value_columns:
            return None
        for column in value_columns:
            used.add(column)
            selected.append(_build_column(worksheet, header_row, column))

    return selected


def _build_column(
    worksheet: Worksheet,
    header_row: int,
    column: int,
) -> _SelectedColumn:
    header = worksheet.cell(row=header_row, column=column).value
    values = _read_column_body(worksheet, header_row + 1, column)
    return _SelectedColumn(header=header, source_column=column, values=values)


def _read_column_body(
    worksheet: Worksheet,
    start_row: int,
    column: int,
) -> list[CellValue]:
    """Read a column downward while value/top-border/bottom-border continues."""

    values: list[CellValue] = []
    row = start_row
    while row <= worksheet.max_row:
        value = worksheet.cell(row=row, column=column).value
        if (
            value is not None
            or _has_top_border(worksheet, row, column)
            or _has_bottom_border(worksheet, row, column)
        ):
            values.append(value)
            row += 1
        else:
            break
    return values


def _find_header_column(
    worksheet: Worksheet,
    header_row: int,
    start_column: int,
    end_column: int,
    expected: CellValue,
    *,
    match_case: bool,
    skip: set[int] | None = None,
) -> int | None:
    target = _normalize_value(expected, match_case=match_case)
    for column in range(start_column, end_column + 1):
        if skip is not None and column in skip:
            continue
        actual = worksheet.cell(row=header_row, column=column).value
        if _normalize_value(actual, match_case=match_case) == target:
            return column
    return None


def _header_contains(value: CellValue, text: str, *, match_case: bool) -> bool:
    haystack = "" if value is None else str(value)
    needle = text
    if not match_case:
        haystack = haystack.casefold()
        needle = needle.casefold()
    return needle in haystack
