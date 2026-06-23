"""Excel-based table format definitions and placeholder extraction."""

from __future__ import annotations

import re
from collections.abc import Iterator, Mapping
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Literal, TYPE_CHECKING

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from openpyxlwings.exceptions import (
    BorderTableNotFoundError,
    BorderTableShapeError,
    FormatDefinitionError,
    FormatMatchError,
    FormatValueError,
)

if TYPE_CHECKING:
    from openpyxlwings.workbook import ExcelWorkbook


CellValue = str | int | float | bool | date | datetime | None
PlaceholderKind = Literal["scalar", "row", "column", "matrix"]

_PLACEHOLDER_RE = re.compile(r"^\s*\{\{(?P<body>.*?)\}\}\s*$")
_CONSTRAINT_RE = re.compile(r'^(contains|equals)\(\s*"([^"]*)"\s*\)$')
_VALUE_TYPES = {"str", "int", "float", "bool", "date", "datetime"}


@dataclass(frozen=True)
class PlaceholderConstraint:
    operator: Literal["contains", "equals"]
    expected: str

    def matches(self, value: CellValue) -> bool:
        actual = "" if value is None else str(value)
        if self.operator == "contains":
            return self.expected in actual
        return actual == self.expected


@dataclass(frozen=True)
class Placeholder:
    """Parsed placeholder from a format cell."""

    expression: str
    kind: PlaceholderKind
    group: str | None
    field: str
    value_type: str | None = None
    constraints: tuple[PlaceholderConstraint, ...] = ()

    @classmethod
    def parse(cls, value: str) -> Placeholder | None:
        match = _PLACEHOLDER_RE.match(value)
        if match is None:
            return None

        expression = match.group("body").strip()
        parts = [part.strip() for part in expression.split("|")]
        path_and_type = parts[0]
        constraints = tuple(_parse_constraint(part) for part in parts[1:])

        path, value_type = _split_value_type(path_and_type)
        segments = path.split(".")
        if not segments or any(not segment for segment in segments):
            raise FormatDefinitionError(f"Invalid placeholder path: {value}")

        if segments[0] == "rows[]":
            if len(segments) != 2:
                raise FormatDefinitionError(f"Row placeholder must have one field: {value}")
            matrix = segments[1].endswith("[]")
            field_name = segments[1][:-2] if matrix else segments[1]
            kind: PlaceholderKind = "matrix" if matrix else "row"
            group = "rows"
        elif segments[0] == "columns[]":
            if len(segments) != 2 or segments[1].endswith("[]"):
                raise FormatDefinitionError(f"Column placeholder must have one field: {value}")
            field_name = segments[1]
            kind = "column"
            group = "columns"
        else:
            if len(segments) != 1 or segments[0].endswith("[]"):
                raise FormatDefinitionError(f"Unsupported placeholder path: {value}")
            field_name = segments[0]
            kind = "scalar"
            group = None

        if not field_name.isidentifier():
            raise FormatDefinitionError(f"Invalid placeholder field name: {field_name}")
        return cls(
            expression=expression,
            kind=kind,
            group=group,
            field=field_name,
            value_type=value_type,
            constraints=constraints,
        )

    def convert(self, value: CellValue, *, source: str) -> CellValue:
        for constraint in self.constraints:
            if not constraint.matches(value):
                raise FormatValueError(
                    f"Value at {source} does not satisfy {constraint.operator}"
                    f"({constraint.expected!r})."
                )
        if self.value_type is None or value is None:
            return value
        try:
            return _convert_value(value, self.value_type)
        except (TypeError, ValueError) as exc:
            raise FormatValueError(
                f"Value at {source} cannot be converted to {self.value_type}: {value!r}"
            ) from exc


@dataclass(frozen=True)
class PatternCell:
    row: int
    column: int
    literal: CellValue = None
    placeholder: Placeholder | None = None

    @property
    def ignored(self) -> bool:
        return self.literal is None and self.placeholder is None


@dataclass(frozen=True)
class TablePattern:
    """Normalized pattern loaded from one format workbook sheet."""

    name: str
    cells: tuple[tuple[PatternCell, ...], ...]
    source_range: str
    row_repeat_index: int | None
    column_repeat_index: int | None
    anchor: PatternCell | None

    @property
    def row_count(self) -> int:
        return len(self.cells)

    @property
    def column_count(self) -> int:
        return len(self.cells[0]) if self.cells else 0

    def iter_cells(self) -> Iterator[PatternCell]:
        for row in self.cells:
            yield from row


@dataclass(frozen=True)
class ExtractedMatch:
    """One table matched and converted from a target workbook."""

    sheet: str
    range: str
    data: dict[str, Any]
    formulas: dict[str, Any]
    source_cells: dict[str, Any]


@dataclass
class ExcelFormat(Mapping[str, TablePattern]):
    """Collection of table patterns loaded from an Excel workbook."""

    path: Path
    patterns: dict[str, TablePattern] = field(default_factory=dict)

    @classmethod
    def load(cls, path: str | Path) -> ExcelFormat:
        format_path = Path(path)
        workbook = load_workbook(format_path, read_only=False, data_only=False)
        try:
            patterns = {
                worksheet.title: _parse_pattern_sheet(worksheet)
                for worksheet in workbook.worksheets
                if _worksheet_has_pattern_content(worksheet)
            }
        finally:
            workbook.close()
        if not patterns:
            raise FormatDefinitionError("The format workbook does not contain any patterns.")
        return cls(path=format_path, patterns=patterns)

    def __getitem__(self, key: str) -> TablePattern:
        return self.patterns[key]

    def __iter__(self) -> Iterator[str]:
        return iter(self.patterns)

    def __len__(self) -> int:
        return len(self.patterns)


def extract_pattern(
    workbook: ExcelWorkbook,
    pattern: TablePattern,
    *,
    sheets: list[str] | None = None,
    ranges: dict[str, str] | None = None,
) -> list[ExtractedMatch]:
    """Find and extract every target table matching ``pattern``."""

    values_book = load_workbook(workbook.path, read_only=False, data_only=True)
    formulas_book = load_workbook(workbook.path, read_only=False, data_only=False)
    try:
        selected_sheets = sheets or (list(ranges) if ranges else list(values_book.sheetnames))
        unknown = [name for name in selected_sheets if name not in values_book.sheetnames]
        if unknown:
            raise FormatMatchError(f"Sheets not found: {', '.join(unknown)}")
        if pattern.anchor is None and not ranges:
            raise FormatMatchError(
                "Patterns without a fixed literal require an explicit ranges argument."
            )

        matches: list[ExtractedMatch] = []
        seen: set[tuple[str, int, int, int, int]] = set()
        for sheet_name in selected_sheets:
            values_sheet = values_book[sheet_name]
            formulas_sheet = formulas_book[sheet_name]
            search_bounds = _search_bounds(values_sheet, None if ranges is None else ranges.get(sheet_name))
            for candidate_row, candidate_column in _candidate_origins(
                values_sheet,
                pattern,
                search_bounds,
            ):
                try:
                    bounds = _detect_candidate_bounds(
                        workbook,
                        values_sheet,
                        candidate_row,
                        candidate_column,
                    )
                    if not _bounds_within(bounds, search_bounds):
                        continue
                    key = (sheet_name, *bounds)
                    if key in seen:
                        continue
                    match = _match_and_extract(
                        pattern,
                        values_sheet,
                        formulas_sheet,
                        bounds,
                    )
                except (BorderTableNotFoundError, BorderTableShapeError, FormatValueError):
                    continue
                seen.add(key)
                matches.append(match)
        return sorted(matches, key=lambda item: (selected_sheets.index(item.sheet), _range_sort_key(item.range)))
    finally:
        values_book.close()
        formulas_book.close()


def _parse_pattern_sheet(worksheet: Worksheet) -> TablePattern:
    min_row, min_column, max_row, max_column = _pattern_bounds(worksheet)
    _validate_pattern_grid(worksheet, min_row, min_column, max_row, max_column)
    rows: list[tuple[PatternCell, ...]] = []
    row_repeat_indexes: set[int] = set()
    column_repeat_indexes: set[int] = set()
    anchor: PatternCell | None = None

    for row_index, excel_row in enumerate(range(min_row, max_row + 1)):
        parsed_row: list[PatternCell] = []
        for column_index, excel_column in enumerate(range(min_column, max_column + 1)):
            value = worksheet.cell(excel_row, excel_column).value
            placeholder = Placeholder.parse(value) if isinstance(value, str) else None
            literal = None if placeholder is not None else value
            cell = PatternCell(row=row_index, column=column_index, literal=literal, placeholder=placeholder)
            parsed_row.append(cell)
            if anchor is None and literal is not None:
                anchor = cell
            if placeholder is not None:
                if placeholder.kind in {"row", "matrix"}:
                    row_repeat_indexes.add(row_index)
                if placeholder.kind in {"column", "matrix"}:
                    column_repeat_indexes.add(column_index)
        rows.append(tuple(parsed_row))

    row_repeat_index = _single_repeat_index(row_repeat_indexes, "row")
    column_repeat_index = _single_repeat_index(column_repeat_indexes, "column")
    if row_repeat_index is not None and row_repeat_index != len(rows) - 1:
        raise FormatDefinitionError("The repeating row prototype must be the last pattern row.")
    if column_repeat_index is not None and column_repeat_index != len(rows[0]) - 1:
        raise FormatDefinitionError("The repeating column prototype must be the last pattern column.")

    _validate_placeholder_positions(rows, row_repeat_index, column_repeat_index)
    source_range = _range_address(min_row, min_column, max_row, max_column)
    return TablePattern(
        name=worksheet.title,
        cells=tuple(rows),
        source_range=source_range,
        row_repeat_index=row_repeat_index,
        column_repeat_index=column_repeat_index,
        anchor=anchor,
    )


def _match_and_extract(
    pattern: TablePattern,
    values_sheet: Worksheet,
    formulas_sheet: Worksheet,
    bounds: tuple[int, int, int, int],
) -> ExtractedMatch:
    start_row, start_column, end_row, end_column = bounds
    target_rows = end_row - start_row + 1
    target_columns = end_column - start_column + 1
    if target_rows < pattern.row_count or target_columns < pattern.column_count:
        raise FormatMatchError("Candidate table is smaller than the format pattern.")
    if pattern.row_repeat_index is None and target_rows != pattern.row_count:
        raise FormatMatchError("Candidate row count does not match the fixed pattern.")
    if pattern.column_repeat_index is None and target_columns != pattern.column_count:
        raise FormatMatchError("Candidate column count does not match the fixed pattern.")

    data: dict[str, Any] = {}
    formulas: dict[str, Any] = {}
    sources: dict[str, Any] = {}
    if pattern.row_repeat_index is not None:
        row_count = target_rows - pattern.row_repeat_index
        data["rows"] = [{} for _ in range(row_count)]
        formulas["rows"] = [{} for _ in range(row_count)]
        sources["rows"] = [{} for _ in range(row_count)]
    if pattern.column_repeat_index is not None:
        column_count = target_columns - pattern.column_repeat_index
        data["columns"] = [{} for _ in range(column_count)]
        formulas["columns"] = [{} for _ in range(column_count)]
        sources["columns"] = [{} for _ in range(column_count)]

    for pattern_cell in pattern.iter_cells():
        for target_row, target_column, row_index, column_index in _mapped_coordinates(
            pattern_cell,
            pattern,
            bounds,
        ):
            value = values_sheet.cell(target_row, target_column).value
            raw_formula = formulas_sheet.cell(target_row, target_column).value
            formula = raw_formula if isinstance(raw_formula, str) and raw_formula.startswith("=") else None
            source = f"{values_sheet.title}!{_cell_address(target_row, target_column)}"
            if pattern_cell.literal is not None:
                if value != pattern_cell.literal:
                    raise FormatMatchError(
                        f"Literal mismatch at {source}: expected {pattern_cell.literal!r}, got {value!r}."
                    )
                continue
            placeholder = pattern_cell.placeholder
            if placeholder is None:
                continue
            converted = placeholder.convert(value, source=source)
            _store_extracted(
                data,
                formulas,
                sources,
                placeholder,
                converted,
                formula,
                source,
                row_index,
                column_index,
            )

    return ExtractedMatch(
        sheet=values_sheet.title,
        range=_range_address(*bounds),
        data=data,
        formulas=formulas,
        source_cells=sources,
    )


def _store_extracted(
    data: dict[str, Any],
    formulas: dict[str, Any],
    sources: dict[str, Any],
    placeholder: Placeholder,
    value: CellValue,
    formula: str | None,
    source: str,
    row_index: int | None,
    column_index: int | None,
) -> None:
    if placeholder.kind == "scalar":
        data[placeholder.field] = value
        formulas[placeholder.field] = formula
        sources[placeholder.field] = source
    elif placeholder.kind == "row":
        _require_index(row_index, "row")
        data["rows"][row_index][placeholder.field] = value
        formulas["rows"][row_index][placeholder.field] = formula
        sources["rows"][row_index][placeholder.field] = source
    elif placeholder.kind == "column":
        _require_index(column_index, "column")
        data["columns"][column_index][placeholder.field] = value
        formulas["columns"][column_index][placeholder.field] = formula
        sources["columns"][column_index][placeholder.field] = source
    else:
        _require_index(row_index, "row")
        _require_index(column_index, "column")
        data_row = data["rows"][row_index]
        formula_row = formulas["rows"][row_index]
        source_row = sources["rows"][row_index]
        data_row.setdefault(placeholder.field, [None] * len(data.get("columns", [])))
        formula_row.setdefault(placeholder.field, [None] * len(data.get("columns", [])))
        source_row.setdefault(placeholder.field, [None] * len(data.get("columns", [])))
        data_row[placeholder.field][column_index] = value
        formula_row[placeholder.field][column_index] = formula
        source_row[placeholder.field][column_index] = source


def _mapped_coordinates(
    cell: PatternCell,
    pattern: TablePattern,
    bounds: tuple[int, int, int, int],
) -> Iterator[tuple[int, int, int | None, int | None]]:
    start_row, start_column, end_row, end_column = bounds
    if pattern.row_repeat_index is not None and cell.row == pattern.row_repeat_index:
        target_rows = range(start_row + cell.row, end_row + 1)
    else:
        target_rows = (start_row + cell.row,)
    if pattern.column_repeat_index is not None and cell.column == pattern.column_repeat_index:
        target_columns = range(start_column + cell.column, end_column + 1)
    else:
        target_columns = (start_column + cell.column,)

    for target_row in target_rows:
        row_index = (
            target_row - (start_row + pattern.row_repeat_index)
            if pattern.row_repeat_index is not None and cell.row == pattern.row_repeat_index
            else None
        )
        for target_column in target_columns:
            column_index = (
                target_column - (start_column + pattern.column_repeat_index)
                if pattern.column_repeat_index is not None and cell.column == pattern.column_repeat_index
                else None
            )
            yield target_row, target_column, row_index, column_index


def _candidate_origins(
    worksheet: Worksheet,
    pattern: TablePattern,
    bounds: tuple[int, int, int, int],
) -> Iterator[tuple[int, int]]:
    min_row, min_column, max_row, max_column = bounds
    if pattern.anchor is None:
        yield min_row, min_column
        return
    anchor = pattern.anchor
    for row in range(min_row, max_row + 1):
        for column in range(min_column, max_column + 1):
            if worksheet.cell(row, column).value != anchor.literal:
                continue
            origin_row = row - anchor.row
            origin_column = column - anchor.column
            if origin_row >= min_row and origin_column >= min_column:
                yield origin_row, origin_column


def _detect_candidate_bounds(
    workbook: ExcelWorkbook,
    worksheet: Worksheet,
    row: int,
    column: int,
) -> tuple[int, int, int, int]:
    from openpyxlwings.border_table import detect_bordered_table

    table = detect_bordered_table(
        workbook,
        worksheet,
        worksheet.title,
        row,
        column,
        header_rows=0,
        header_columns=0,
    )
    return table.start_row, table.start_column, table.end_row, table.end_column


def _pattern_bounds(worksheet: Worksheet) -> tuple[int, int, int, int]:
    coordinates = [
        (cell.row, cell.column)
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value is not None or _cell_has_border(cell)
    ]
    if not coordinates:
        raise FormatDefinitionError(f"Pattern sheet is empty: {worksheet.title}")
    rows, columns = zip(*coordinates, strict=True)
    return min(rows), min(columns), max(rows), max(columns)


def _worksheet_has_pattern_content(worksheet: Worksheet) -> bool:
    return any(
        cell.value is not None or _cell_has_border(cell)
        for row in worksheet.iter_rows()
        for cell in row
    )


def _validate_placeholder_positions(
    rows: list[tuple[PatternCell, ...]],
    row_repeat_index: int | None,
    column_repeat_index: int | None,
) -> None:
    seen: set[tuple[PlaceholderKind, str]] = set()
    for row in rows:
        for cell in row:
            placeholder = cell.placeholder
            if placeholder is None:
                continue
            key = (placeholder.kind, placeholder.field)
            if key in seen and placeholder.kind != "matrix":
                raise FormatDefinitionError(f"Duplicate placeholder: {placeholder.expression}")
            seen.add(key)
            if placeholder.kind in {"row", "matrix"} and cell.row != row_repeat_index:
                raise FormatDefinitionError("Row placeholders must be in the repeating row prototype.")
            if placeholder.kind in {"column", "matrix"} and cell.column != column_repeat_index:
                raise FormatDefinitionError("Column placeholders must be in the repeating column prototype.")
            if placeholder.kind == "scalar" and (
                cell.row == row_repeat_index or cell.column == column_repeat_index
            ):
                raise FormatDefinitionError("Scalar placeholders must be outside repeating prototypes.")
            if placeholder.kind == "row" and cell.column == column_repeat_index:
                raise FormatDefinitionError("Row placeholders must be in a fixed column.")
            if placeholder.kind == "column" and cell.row == row_repeat_index:
                raise FormatDefinitionError("Column placeholders must be in a fixed row.")


def _validate_pattern_grid(
    worksheet: Worksheet,
    start_row: int,
    start_column: int,
    end_row: int,
    end_column: int,
) -> None:
    for merged_range in worksheet.merged_cells.ranges:
        if not (
            merged_range.max_row < start_row
            or merged_range.min_row > end_row
            or merged_range.max_col < start_column
            or merged_range.min_col > end_column
        ):
            raise FormatDefinitionError("Merged cells are not supported in format patterns.")

    for row in range(start_row, end_row + 1):
        for column in range(start_column, end_column + 1):
            cell = worksheet.cell(row, column)
            if row == start_row and not _has_side(cell.border.top):
                raise FormatDefinitionError("Pattern top border is incomplete.")
            if row == end_row and not _has_side(cell.border.bottom):
                raise FormatDefinitionError("Pattern bottom border is incomplete.")
            if column == start_column and not _has_side(cell.border.left):
                raise FormatDefinitionError("Pattern left border is incomplete.")
            if column == end_column and not _has_side(cell.border.right):
                raise FormatDefinitionError("Pattern right border is incomplete.")
            if row < end_row and not (
                _has_side(cell.border.bottom)
                or _has_side(worksheet.cell(row + 1, column).border.top)
            ):
                raise FormatDefinitionError("Pattern has a missing horizontal border.")
            if column < end_column and not (
                _has_side(cell.border.right)
                or _has_side(worksheet.cell(row, column + 1).border.left)
            ):
                raise FormatDefinitionError("Pattern has a missing vertical border.")


def _split_value_type(path_and_type: str) -> tuple[str, str | None]:
    if ":" not in path_and_type:
        return path_and_type.strip(), None
    path, value_type = path_and_type.rsplit(":", 1)
    value_type = value_type.strip()
    if value_type not in _VALUE_TYPES:
        raise FormatDefinitionError(f"Unsupported placeholder type: {value_type}")
    return path.strip(), value_type


def _parse_constraint(value: str) -> PlaceholderConstraint:
    match = _CONSTRAINT_RE.match(value)
    if match is None:
        raise FormatDefinitionError(f"Unsupported placeholder constraint: {value}")
    return PlaceholderConstraint(operator=match.group(1), expected=match.group(2))


def _convert_value(value: CellValue, value_type: str) -> CellValue:
    if value_type == "str":
        return str(value)
    if value_type == "int":
        if isinstance(value, bool):
            raise ValueError
        number = int(value)
        if isinstance(value, float) and value != number:
            raise ValueError
        return number
    if value_type == "float":
        if isinstance(value, bool):
            raise ValueError
        return float(value)
    if value_type == "bool":
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            normalized = value.strip().casefold()
            if normalized in {"true", "yes", "1"}:
                return True
            if normalized in {"false", "no", "0"}:
                return False
        if value in {0, 1}:
            return bool(value)
        raise ValueError
    if value_type == "date":
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return date.fromisoformat(str(value))
    if value_type == "datetime":
        if isinstance(value, datetime):
            return value
        return datetime.fromisoformat(str(value))
    raise ValueError


def _search_bounds(worksheet: Worksheet, address: str | None) -> tuple[int, int, int, int]:
    if address is None:
        return 1, 1, worksheet.max_row, worksheet.max_column
    min_column, min_row, max_column, max_row = range_boundaries(address)
    return min_row, min_column, max_row, max_column


def _single_repeat_index(indexes: set[int], axis: str) -> int | None:
    if len(indexes) > 1:
        raise FormatDefinitionError(f"Only one repeating {axis} prototype is supported.")
    return next(iter(indexes), None)


def _cell_has_border(cell: Cell) -> bool:
    return any(_has_side(side) for side in (cell.border.top, cell.border.bottom, cell.border.left, cell.border.right))


def _has_side(side: object) -> bool:
    return bool(getattr(side, "style", None))


def _cell_address(row: int, column: int) -> str:
    return f"{get_column_letter(column)}{row}"


def _range_address(start_row: int, start_column: int, end_row: int, end_column: int) -> str:
    return f"{_cell_address(start_row, start_column)}:{_cell_address(end_row, end_column)}"


def _range_sort_key(address: str) -> tuple[int, int]:
    min_column, min_row, _max_column, _max_row = range_boundaries(address)
    return min_row, min_column


def _bounds_within(
    bounds: tuple[int, int, int, int],
    container: tuple[int, int, int, int],
) -> bool:
    start_row, start_column, end_row, end_column = bounds
    min_row, min_column, max_row, max_column = container
    return (
        start_row >= min_row
        and start_column >= min_column
        and end_row <= max_row
        and end_column <= max_column
    )


def _require_index(value: int | None, axis: str) -> None:
    if value is None:
        raise FormatMatchError(f"Missing {axis} repetition index.")
